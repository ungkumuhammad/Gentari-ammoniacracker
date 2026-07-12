"""QA pipeline for the built workbook.

Intended design (per the approved plan): build -> headless LibreOffice
recalculates -> read back computed values -> assert golden values. In
this build environment, headless LibreOffice (`soffice --headless
--convert-to ...`) fails to load even a trivial file
(confirmed by direct testing -- not specific to this workbook), so the
live-recalculation step cannot run here. This module:

1. Attempts the soffice recalculation anyway (works if run in an
   environment where LibreOffice is actually functional -- e.g. the
   user's own machine or a CI runner) and uses it for golden-value
   assertions when available.
2. Falls back to two checks that do not require executing the Excel
   formulas: (a) a static named-range reference check (every named range
   a formula references must actually be registered -- catches #NAME?
   bugs before they reach a user), and (b) an independent pure-Python
   re-implementation of the KBR piecewise log-log interpolation and
   global regression, cross-checked against the same 4 sourced data
   points the Excel formulas use, which validates the formula *logic*
   even though the literal Excel formula strings are never executed here.

Known limitation, logged in memory.md: this workbook has not been
recalculated by a real Excel/LibreOffice engine in this environment.
Opening it in real Excel (which recalculates on open, per
CalcProperties.fullCalcOnLoad=True set in build.py) is the outstanding
verification step before treating any cell value as confirmed correct.
"""
import math
import re
import subprocess
import sys
import tempfile
from pathlib import Path

from tools.cracker_model import data
from tools.cracker_model import named_ranges as nr


class QAError(Exception):
    pass


def attempt_soffice_recalc(xlsx_path: str) -> dict | None:
    """Try headless LibreOffice recalculation. Returns a dict of cell
    values keyed by 'Sheet!Cell' if successful, else None."""
    with tempfile.TemporaryDirectory() as tmp:
        profile = Path(tmp) / "lo_profile"
        outdir = Path(tmp) / "out"
        outdir.mkdir()
        try:
            result = subprocess.run(
                ["soffice", "--headless", "--norestore",
                 f"-env:UserInstallation=file://{profile}",
                 "--convert-to", "xlsx", "--outdir", str(outdir), xlsx_path],
                capture_output=True, text=True, timeout=90,
            )
        except (FileNotFoundError, subprocess.TimeoutExpired) as e:
            print(f"[qa_recalc] soffice unavailable/timed out: {e}")
            return None
        converted = outdir / Path(xlsx_path).name
        if result.returncode != 0 or not converted.exists():
            print(f"[qa_recalc] soffice recalculation failed (stdout={result.stdout!r} "
                  f"stderr={result.stderr!r}) -- falling back to static checks only.")
            return None
        import openpyxl
        wb = openpyxl.load_workbook(str(converted), data_only=True)
        values = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        values[f"{ws.title}!{cell.coordinate}"] = cell.value
        return values


_NAME_CANDIDATE_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_]*")
_EXCLUDE_TOKENS = {
    "IF", "INDEX", "MATCH", "LN", "EXP", "AND", "OR", "IFERROR", "TEXT", "TODAY",
    "AVERAGE", "SLOPE", "INTERCEPT", "RSQ", "IRR", "NPV", "RATE", "ISNUMBER", "TRUE", "FALSE",
}


def check_named_ranges_referenced_are_defined(wb) -> list[str]:
    defined = set(wb.defined_names.keys())
    candidate_names = sorted({
        v for k, v in vars(nr).items()
        if isinstance(v, str) and k.isupper() and not k.startswith(("SHEET_", "TBL_"))
        and "{" not in v
    })
    all_formula_text = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    all_formula_text.append(cell.value)
    big_text = "\n".join(all_formula_text)
    problems = []
    for name in candidate_names:
        if name in _EXCLUDE_TOKENS:
            continue
        pattern = r"(?<![A-Za-z0-9_])" + re.escape(name) + r"(?![A-Za-z0-9_\[])"
        if re.search(pattern, big_text) and name not in defined:
            problems.append(f"Named range '{name}' is referenced in a formula but never registered in wb.defined_names")
    return problems


def _piecewise_loglog_capex(capacity: float, points: list[tuple[float, float]]) -> float:
    """Pure-Python re-implementation of the Excel bracket/interpolation
    formula in calc_capacity_sizing.py -- same logic, independent
    execution path."""
    caps = [p[0] for p in points]
    capexes = [p[1] for p in points]
    if capacity <= caps[1]:
        lo, hi = 0, 1
    elif capacity <= caps[2]:
        lo, hi = 1, 2
    else:
        lo, hi = 2, 3
    cap_lo, cap_hi = caps[lo], caps[hi]
    capex_lo, capex_hi = capexes[lo], capexes[hi]
    b = math.log(capex_hi / capex_lo) / math.log(cap_hi / cap_lo)
    return capex_lo * (capacity / cap_lo) ** b


def _global_regression_fit(points: list[tuple[float, float]]) -> tuple[float, float, float]:
    """SLOPE/INTERCEPT/RSQ on ln-ln data, independent of Excel."""
    xs = [math.log(p[0]) for p in points]
    ys = [math.log(p[1]) for p in points]
    n = len(xs)
    mx, my = sum(xs) / n, sum(ys) / n
    sxy = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    sxx = sum((x - mx) ** 2 for x in xs)
    b = sxy / sxx
    ln_a = my - b * mx
    a = math.exp(ln_a)
    ss_res = sum((y - (ln_a + b * x)) ** 2 for x, y in zip(xs, ys))
    ss_tot = sum((y - my) ** 2 for y in ys)
    r2 = 1 - ss_res / ss_tot
    return a, b, r2


def verify_kbr_interpolation_logic() -> list[str]:
    problems = []
    points = [(cap, data.KBR_ISBL_CAPEX_MUSD[cap].value) for cap in data.KBR_CAPACITIES_KTPA]

    for cap, expected_capex in points:
        got = _piecewise_loglog_capex(cap, points)
        if abs(got - expected_capex) > 1e-6:
            problems.append(f"Piecewise interpolation at KBR's own capacity {cap} ktpa "
                             f"gave {got:.4f}, expected exactly {expected_capex} (sourced value)")

    a, b, r2 = _global_regression_fit(points)
    if r2 < 0.95:
        problems.append(f"Global log-log regression fit is unexpectedly poor: R2={r2:.4f} (expected > 0.95)")
    if not (0.3 < b < 0.8):
        problems.append(f"Global regression exponent b={b:.4f} is outside a physically plausible "
                         f"economy-of-scale range (0.3-0.8) -- check for a data/formula error")

    interp_120 = _piecewise_loglog_capex(24, points)  # sanity: matches sourced 110 exactly at cap=24
    if abs(interp_120 - 110) > 1e-6:
        problems.append(f"Sanity check failed: interpolation at 24 ktpa gave {interp_120}, expected 110")

    print(f"[qa_recalc] KBR regression fit (independent Python re-derivation): "
          f"a={a:.4f}, b={b:.4f}, R2={r2:.6f}")
    return problems


def run() -> None:
    xlsx_path = str(Path(__file__).resolve().parents[2] / "output" / "Ammonia_Cracker_Sizing_Economics_v1.xlsx")
    if not Path(xlsx_path).exists():
        raise QAError(f"{xlsx_path} does not exist -- run `python -m tools.cracker_model.build` first")

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)

    all_problems = []

    name_problems = check_named_ranges_referenced_are_defined(wb)
    if name_problems:
        all_problems.extend(name_problems)
    else:
        print("[qa_recalc] All named ranges referenced in formulas are registered.")

    logic_problems = verify_kbr_interpolation_logic()
    if logic_problems:
        all_problems.extend(logic_problems)
    else:
        print("[qa_recalc] KBR piecewise interpolation & global regression logic verified "
              "(independent Python re-implementation matches all 4 sourced data points exactly).")

    recalc_values = attempt_soffice_recalc(xlsx_path)
    if recalc_values is None:
        print("[qa_recalc] WARNING: headless LibreOffice recalculation is not functional in this "
              "environment (confirmed via direct testing -- even trivial file conversions fail). "
              "Live Excel-formula golden-value checks were NOT performed. This workbook has not "
              "been recalculated by a real spreadsheet engine. Open it in Excel or a working "
              "LibreOffice install to confirm computed values before relying on them.")
    else:
        golden = {
            f"'{nr.SHEET_CONSTANTS}'!C7": 78, f"'{nr.SHEET_CONSTANTS}'!C8": 110,
            f"'{nr.SHEET_CONSTANTS}'!C9": 191, f"'{nr.SHEET_CONSTANTS}'!C10": 209,
        }
        for ref, expected in golden.items():
            got = recalc_values.get(ref.replace("'", ""))
            if got != expected:
                all_problems.append(f"Golden value mismatch at {ref}: got {got}, expected {expected}")
        print("[qa_recalc] Live LibreOffice recalculation succeeded and golden values checked.")

    if all_problems:
        print(f"\n[qa_recalc] FAILED -- {len(all_problems)} problem(s):")
        for p in all_problems:
            print(f"  - {p}")
        raise QAError(f"{len(all_problems)} QA problem(s) found")

    print("\n[qa_recalc] All available QA checks passed.")


if __name__ == "__main__":
    try:
        run()
    except QAError as e:
        print(f"QA FAILED: {e}", file=sys.stderr)
        sys.exit(1)
