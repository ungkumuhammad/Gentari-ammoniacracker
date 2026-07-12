"""Entry point: python -m tools.cracker_model.build

Orchestrates every sheet builder, wires cross-sheet row references, sets
tab order to match named_ranges.SHEET_ORDER, applies admin-protection to
Constants/Settings (documented deterrent password per mdlguideline.md
§12), and saves the workbook to output/.
"""
import os

import openpyxl
from openpyxl.workbook.properties import CalcProperties

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import validate_data
from tools.cracker_model.sheets import (
    admin_constants, calc_capacity_sizing, calc_capex_isbl_osbl, calc_carbon_intensity,
    calc_cashflow_irr, calc_opex, calc_regional_factor, calc_tolling, comparison,
    cover, dashboard, guide, inputs, report, settings,
)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Ammonia_Cracker_Sizing_Economics_v1.xlsx")


def build_workbook() -> openpyxl.Workbook:
    validate_data.run()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True)

    cover.build(wb)
    guide.build(wb)
    inputs.build(wb)
    admin_constants.build(wb)
    ws_cs, cs_rows = calc_capacity_sizing.build(wb)
    ws_capex, capex_rows = calc_capex_isbl_osbl.build(wb, cs_rows)
    ws_opex, opex_rows = calc_opex.build(wb, capex_rows)
    ws_toll, toll_rows = calc_tolling.build(wb)
    ws_reg, reg_rows = calc_regional_factor.build(wb)
    ws_ci, ci_rows = calc_carbon_intensity.build(wb)
    ws_cf, cf_rows = calc_cashflow_irr.build(wb)
    dashboard.build(wb, ci_rows, cs_rows, toll_rows, reg_rows, cf_rows)
    comparison.build(wb, cs_rows, capex_rows, opex_rows, ci_rows)
    report.build(wb)
    settings.build(wb)

    # Reorder tabs to match the intended navigation order (Constants/
    # Settings tucked away at the end -- build order above was chosen for
    # formula-wiring convenience, not display order).
    wb._sheets = [wb[name] for name in nr.SHEET_ORDER]

    # Admin protection: documented deterrent password, per mdlguideline.md
    # §12 and the approved plan's sign-off (protects against accidental
    # edits, not a determined editor).
    for sheet_name in nr.PROTECTED_SHEETS:
        ws = wb[sheet_name]
        ws.protection.sheet = True
        ws.protection.set_password(settings.ADMIN_PASSWORD)
        ws.protection.enable()

    wb.active = wb.sheetnames.index(nr.SHEET_COVER)
    return wb


def main() -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Named ranges: {len(wb.defined_names)}")


if __name__ == "__main__":
    main()
