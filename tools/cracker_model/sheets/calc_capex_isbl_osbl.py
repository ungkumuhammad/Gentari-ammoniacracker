"""Calc_CAPEX_ISBL_OSBL -- ISBL/OSBL split and region-adjusted CAPEX total
per licensor, feeding CAPEX_Total_Own for Calc_CashFlow_IRR."""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

LIC = nr.LICENSOR_SELECTED
SHEET = nr.SHEET_CALC_CAPACITY_SIZING


def build(wb, capacity_sizing_rows: dict) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_CALC_CAPEX_ISBL_OSBL)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["F"].width = 55

    kbr_capex_ref = f"'{SHEET}'!$D${capacity_sizing_rows['kbr_capex_row']}"
    duiker_ref = f"'{SHEET}'!$D${capacity_sizing_rows['duiker_row']}"
    topsoe_ref = f"'{SHEET}'!$D${capacity_sizing_rows['topsoe_row']}"

    r = 2
    ws.cell(r, 2, "Calc_CAPEX_ISBL_OSBL")
    st.style_title(ws.cell(r, 2))
    r += 2

    ws.cell(r, 2, "KBR")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "KBR_ISBL_CAPEX_MUSD")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"={kbr_capex_ref}")
    st.style_calculated(ws.cell(r, 4))
    kbr_isbl_row = r
    r += 1
    ws.cell(r, 2, "KBR_OSBL_CAPEX_MUSD (ASSUMPTION)")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"=D{kbr_isbl_row}*{nr.KBR_OSBL_PCT_ASSUMPTION}")
    st.style_assumption_flag(ws.cell(r, 4))
    kbr_osbl_row = r
    r += 1
    ws.cell(r, 2, "KBR_Total_CAPEX_MUSD (pre-region)")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"=D{kbr_isbl_row}+D{kbr_osbl_row}")
    st.style_calculated(ws.cell(r, 4))
    kbr_total_pre_region_row = r
    r += 1
    ws.cell(r, 2, "KBR_Total_CAPEX_MUSD (region-adjusted)")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"=D{kbr_total_pre_region_row}*{nr.REGIONAL_FACTOR_SELECTED}")
    st.style_calculated(ws.cell(r, 4))
    kbr_total_row = r
    r += 2

    ws.cell(r, 2, "Duiker")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "Duiker_Total_Installed_Cost_MUSD (no ISBL/OSBL split)")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"={duiker_ref}")
    st.style_calculated(ws.cell(r, 4))
    duiker_total_pre_row = r
    r += 1
    ws.cell(r, 2, "Duiker_Total_CAPEX_MUSD (region-adjusted)")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f'=IFERROR(D{duiker_total_pre_row}*{nr.REGIONAL_FACTOR_SELECTED},D{duiker_total_pre_row})')
    st.style_calculated(ws.cell(r, 4))
    duiker_total_row = r
    r += 2

    ws.cell(r, 2, "Topsoe")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "Topsoe_ISBL_CAPEX_MUSD (OSBL not available -- total understates true installed cost)")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"={topsoe_ref}")
    st.style_calculated(ws.cell(r, 4))
    topsoe_total_pre_row = r
    r += 1
    ws.cell(r, 2, "Topsoe_Total_CAPEX_MUSD (region-adjusted, ISBL-only basis)")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f'=IFERROR(D{topsoe_total_pre_row}*{nr.REGIONAL_FACTOR_SELECTED},D{topsoe_total_pre_row})')
    st.style_calculated(ws.cell(r, 4))
    topsoe_total_row = r
    r += 2

    ws.cell(r, 2, "Casale / Technip")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "Casale_Total_CAPEX_MUSD")
    st.style_label(ws.cell(r, 2))
    st.style_na_flag(ws.cell(r, 4, "N/A -- technical proposal only, no cost data"))
    r += 1
    ws.cell(r, 2, "Technip_Total_CAPEX_MUSD")
    st.style_label(ws.cell(r, 2))
    st.style_na_flag(ws.cell(r, 4, "N/A -- not available at current maturity"))
    r += 2

    ws.cell(r, 2, "Selected Output (Own & Operate mode)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "CAPEX_Total_Own (MUSD, region-adjusted)")
    st.style_label(ws.cell(r, 2))
    formula = (
        f'=IF({LIC}="KBR",D{kbr_total_row},'
        f'IF({LIC}="Duiker",D{duiker_total_row},'
        f'IF({LIC}="Topsoe",D{topsoe_total_row},'
        f'IF({LIC}="Casale","N/A -- technical proposal only, no cost data",'
        f'"N/A -- not available at current maturity"))))'
    )
    ws.cell(r, 4, formula)
    st.style_calculated(ws.cell(r, 4))
    nr.register(wb, nr.CAPEX_TOTAL_OWN, f"'{nr.SHEET_CALC_CAPEX_ISBL_OSBL}'!$D${r}")
    r += 1

    rows = {
        "kbr_isbl_row": kbr_isbl_row,
        "kbr_osbl_row": kbr_osbl_row,
        "kbr_total_pre_region_row": kbr_total_pre_region_row,
        "kbr_total_row": kbr_total_row,
        "duiker_total_row": duiker_total_row,
        "topsoe_total_row": topsoe_total_row,
    }
    return ws, rows
