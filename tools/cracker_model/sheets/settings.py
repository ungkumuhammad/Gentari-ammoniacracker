"""Settings sheet -- version history, revision notes, protection password
reference. mdlguideline.md §12: documented deterrent password (not real
access control), per user sign-off recorded in the approved plan."""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

ADMIN_PASSWORD = "Gentari-TCOE-2026"  # documented deterrent, not secret -- see note below


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_SETTINGS)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 70

    r = 2
    ws.cell(r, 2, "Settings")
    st.style_title(ws.cell(r, 2))
    r += 2

    ws.cell(r, 2, "Admin protection password")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 3, ADMIN_PASSWORD)
    r += 1
    ws.cell(r, 2, "")
    ws.cell(r, 3, "Documented deterrent, per mdlguideline.md §12: protects the Constants "
                   "sheet against accidental edits during normal use, not a determined "
                   "editor (Excel's native sheet-password hash is not real access "
                   "control). Same password for anyone who needs to edit Constants.")
    ws.cell(r, 3).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
    ws.cell(r, 3).alignment = st.Alignment(wrap_text=True, vertical="top")
    r += 2

    ws.cell(r, 2, "Version history")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    r += 1
    headers = ["Version", "Date", "Change"]
    for i, h in enumerate(headers):
        ws.cell(r, 2 + i, h)
        st.style_label(ws.cell(r, 2 + i))
    r += 1
    ws.cell(r, 2, "1.0")
    ws.cell(r, 3, "2026-07-09 -- Initial build per mdlguideline.md and the approved "
                   "capacity-sizing/economics plan. Known v1 deviation: classic "
                   "INDEX/MATCH/IFERROR formulas used instead of mdlguideline.md §7's "
                   "preferred LET()/LAMBDA() -- see mdlguideline.md §7 and memory.md.")

    return ws
