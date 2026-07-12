"""Dashboard -- executive summary: KPI cards, headline CAPEX/tariff + IRR,
RFNBO/KEEI screening, warnings panel pulling live from Calc_ sheet flags."""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

LIC = nr.LICENSOR_SELECTED
CAP = nr.REQUIRED_H2_CAPACITY_KTPA
MODE = nr.COMMERCIAL_MODE
FUEL = nr.CRACKER_FUEL_MODE


def build(wb, ci_rows: dict, cs_rows: dict, tolling_rows: dict,
          regional_rows: dict, cashflow_rows: dict) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_DASHBOARD)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["F"].width = 60

    ci_sheet = nr.SHEET_CALC_CARBON_INTENSITY

    r = 2
    ws.cell(r, 2, "Dashboard")
    st.style_title(ws.cell(r, 2))
    r += 1
    ws.cell(r, 2, f'=({LIC}&" | "&{CAP}&" ktpa H2 | "&{MODE}&" | "&{FUEL})')
    ws.cell(r, 2).font = st.body_font(italic=True, bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 2

    def put(label, formula, note="", style=st.style_calculated):
        nonlocal r
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        c = ws.cell(r, 4, formula)
        style(c)
        if note:
            ws.cell(r, 6, note)
            ws.cell(r, 6).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
        this_row = r
        r += 1
        return this_row

    ws.cell(r, 2, "Headline Economics")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    put("CAPEX_Total_Own (MUSD, Own & Operate mode)", f"={nr.CAPEX_TOTAL_OWN}")
    put("AnnualOPEX_Own (MUSD/yr, Own & Operate mode)", f"={nr.ANNUAL_OPEX_OWN}")
    put("AnnualTollingFee_Tolling (MUSD/yr, Tolling mode)", f"={nr.ANNUAL_TOLLING_FEE}")
    put("IRR_Result (unlevered)", f"={nr.IRR_RESULT}",
        "No debt/tax/depreciation modeled -- see Calc_CashFlow_IRR")
    put("NPV_Result (MUSD, at DiscountRate_pct)", f"={nr.NPV_RESULT}")
    r += 1

    ws.cell(r, 2, "Key Performance Indicators")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    put("Direct CI (kgCO2/kgH2)", f"='{ci_sheet}'!$D${ci_rows['selected_ci_row']}")
    put("Direct CI (gCO2e/MJ)", f"='{ci_sheet}'!$D${ci_rows['ci_gco2_mj_row']}")
    r += 1

    ws.cell(r, 2, "RFNBO / Korea KEEI Screening")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    rfnbo_disp = put("EU RFNBO", f"='{ci_sheet}'!$D${ci_rows['rfnbo_row']}")
    ws.row_dimensions[rfnbo_disp].height = 45
    ws.cell(rfnbo_disp, 4).alignment = st.Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=rfnbo_disp, start_column=4, end_row=rfnbo_disp, end_column=8)
    keei_disp = put("Korea KEEI", f"='{ci_sheet}'!$D${ci_rows['keei_row']}")
    ws.row_dimensions[keei_disp].height = 45
    ws.cell(keei_disp, 4).alignment = st.Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=keei_disp, start_column=4, end_row=keei_disp, end_column=8)
    r += 1

    ws.cell(r, 2, "Warnings & Flags (live from Calc_ sheets)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    warning_rows = [
        ("CAPEX extrapolation", f"='{nr.SHEET_CALC_CAPACITY_SIZING}'!$D${cs_rows['extrap_flag_row']}"),
        ("Project life vs. licensor basis", f"='{nr.SHEET_CALC_CASHFLOW_IRR}'!$D${cashflow_rows['life_flag_row']}"),
        ("Tolling capacity constraint", f"='{nr.SHEET_CALC_TOLLING}'!$D${tolling_rows['warn_row']}"),
        ("Regional factor status", f"='{nr.SHEET_CALC_REGIONAL_FACTOR}'!$D${regional_rows['status_row']}"),
    ]
    for label, formula in warning_rows:
        put(label, formula, style=st.style_extrapolation_flag)
    r += 1

    return ws
