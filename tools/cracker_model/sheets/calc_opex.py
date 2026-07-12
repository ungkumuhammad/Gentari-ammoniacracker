"""Calc_OPEX -- KBR OPEX interpolated by capacity bracket (fuel-mode aware)
and ammonia price; Duiker base OPEX + license fees; Casale/Topsoe/Technip
N/A. Feeds AnnualOPEX_Own for Calc_CashFlow_IRR.

Scope limitation (v1, logged in memory.md): unlike CAPEX, OPEX is NOT
regression-extrapolated outside KBR's 12-80 ktpa range -- OPEX depends on
capacity, fuel mode, and ammonia price simultaneously, and a robust
multi-dimensional regression was judged disproportionate for v1. Outside
that range, OPEX shows N/A even though CAPEX (Calc_CapacitySizing) shows a
regression estimate.
"""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

LIC = nr.LICENSOR_SELECTED
CAP = nr.REQUIRED_H2_CAPACITY_KTPA
FUEL = nr.CRACKER_FUEL_MODE
NG100_LABEL = "100% Natural Gas"


def build(wb, capex_rows: dict) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_CALC_OPEX)
    kbr_total_pre_region_ref = f"'{nr.SHEET_CALC_CAPEX_ISBL_OSBL}'!$D${capex_rows['kbr_total_pre_region_row']}"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["F"].width = 55

    r = 2
    ws.cell(r, 2, "Calc_OPEX")
    st.style_title(ws.cell(r, 2))
    r += 2

    ws.cell(r, 2, "KBR -- Capacity Bracket (fuel-mode aware)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1

    def put(label, formula, note=""):
        nonlocal r
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        c = ws.cell(r, 4, formula)
        st.style_calculated(c)
        if note:
            ws.cell(r, 6, note)
            ws.cell(r, 6).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
        this_row = r
        r += 1
        return this_row

    is_ng100_row = put("FuelMode_Is_NG100", f'=({FUEL}="{NG100_LABEL}")')
    lo_cap_row = put(
        "OPEX_BracketLo_Cap",
        f"=IF(D{is_ng100_row},IF({CAP}<={nr.KBR_CAP2},{nr.KBR_CAP1},IF({CAP}<={nr.KBR_CAP3},{nr.KBR_CAP2},{nr.KBR_CAP3})),{nr.KBR_CAP1})",
        "NG100: 4-point bracket (12/24/68/80). Non-NG100: fixed 12-80 (only 2 points have data)",
    )
    hi_cap_row = put(
        "OPEX_BracketHi_Cap",
        f"=IF(D{is_ng100_row},IF({CAP}<={nr.KBR_CAP2},{nr.KBR_CAP2},IF({CAP}<={nr.KBR_CAP3},{nr.KBR_CAP3},{nr.KBR_CAP4})),{nr.KBR_CAP4})",
    )
    in_range_row = put("InRange_12to80", f"=AND({CAP}>={nr.KBR_CAP1},{CAP}<={nr.KBR_CAP4})")
    interp_basis_row = put(
        "Interp_Basis_Note",
        f'=IF(D{is_ng100_row},"Exact/interpolated across 4 sourced NG100 points",'
        f'"Interpolated between 12 & 80 ktpa only -- 24/68 ktpa have no non-NG100 data in KBR\'s package")',
    )
    r += 1

    ws.cell(r, 2, "KBR -- OPEX Lookup & Price Interpolation at Bracket Points")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1

    def opex_block(bracket_cap_row, label_prefix):
        nonlocal r
        key_row = put(
            f"{label_prefix}_LookupKey",
            f"=D{bracket_cap_row}&\"|\"&{FUEL}",
        )
        o500_row = put(
            f"{label_prefix}_OPEX_500",
            f"=IFERROR(INDEX(tblKBRCases[OPEX_500],MATCH(D{key_row},tblKBRCases[LookupKey],0)),\"N/A\")",
        )
        o950_row = put(
            f"{label_prefix}_OPEX_950",
            f"=IFERROR(INDEX(tblKBRCases[OPEX_950],MATCH(D{key_row},tblKBRCases[LookupKey],0)),\"N/A\")",
        )
        o1100_row = put(
            f"{label_prefix}_OPEX_1100",
            f"=IFERROR(INDEX(tblKBRCases[OPEX_1100],MATCH(D{key_row},tblKBRCases[LookupKey],0)),\"N/A\")",
        )
        price_interp_row = put(
            f"{label_prefix}_OPEX_at_AmmoniaPrice",
            f"=IF({nr.AMMONIA_PRICE_USD_PER_T}<=950,"
            f"D{o500_row}+({nr.AMMONIA_PRICE_USD_PER_T}-500)*(D{o950_row}-D{o500_row})/(950-500),"
            f"D{o950_row}+({nr.AMMONIA_PRICE_USD_PER_T}-950)*(D{o1100_row}-D{o950_row})/(1100-950))",
            "Piecewise linear across KBR's 500/950/1100 USD-per-t data points",
        )
        return price_interp_row

    lo_price_row = opex_block(lo_cap_row, "Lo")
    hi_price_row = opex_block(hi_cap_row, "Hi")
    r += 1

    ws.cell(r, 2, "KBR -- Capacity Interpolation (log-log) & Final Selection")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    b_row = put(
        "OPEX_SegmentExponent_b",
        f"=IF(D{lo_cap_row}=D{hi_cap_row},0,LN(D{hi_price_row}/D{lo_price_row})/LN(D{hi_cap_row}/D{lo_cap_row}))",
    )
    kbr_opex_interp_row = put(
        "KBR_OPEX_MUSD_per_yr (interpolated)",
        f"=D{lo_price_row}*({CAP}/D{lo_cap_row})^D{b_row}",
    )
    kbr_opex_final_row = put(
        "KBR_OPEX_MUSD_per_yr (final)",
        f"=IF(D{in_range_row},D{kbr_opex_interp_row},\"N/A -- outside KBR's 12-80 ktpa OPEX-sourced range (CAPEX regression-extrapolates; OPEX does not in v1)\")",
    )
    r += 1

    ws.cell(r, 2, "LCOH Cross-Check (audit -- recomputed from this model's own CAPEX+OPEX)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "Equation: LCOH = (CRF*CAPEX_MUSD + OPEX_MUSD_yr) / Capacity_ktpa, "
                   "CRF = r(1+r)^n / ((1+r)^n - 1)")
    ws.cell(r, 2).font = st.body_font(italic=True)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    crf_row = put(
        "CRF (using KBR's own 8% rate & 25-yr life)",
        f"=({nr.KBR_LCOH_DISCOUNT_RATE}/100)*(1+{nr.KBR_LCOH_DISCOUNT_RATE}/100)^{nr.KBR_PROJECT_LIFE_YR}/"
        f"((1+{nr.KBR_LCOH_DISCOUNT_RATE}/100)^{nr.KBR_PROJECT_LIFE_YR}-1)",
    )
    lcoh_model_row = put(
        "LCOH_Model_USD_per_kgH2",
        f"=IF(D{in_range_row},"
        f"(D{crf_row}*{kbr_total_pre_region_ref}+D{kbr_opex_final_row})/{CAP},"
        f"\"N/A\")",
        "Cross-check uses un-region-adjusted KBR_Total_CAPEX_MUSD to match KBR's own basis",
    )
    lcoh_kbr_lookup_row = put(
        "LCOH_KBR_Quoted_USD_per_kgH2 (at bracket, price-interpolated)",
        f"=IF(D{in_range_row},"
        f"IFERROR(INDEX(tblKBRCases[LCOH_500],MATCH(D{lo_cap_row}&\"|\"&{FUEL},tblKBRCases[LookupKey],0))"
        f"+({nr.AMMONIA_PRICE_USD_PER_T}-500)*(INDEX(tblKBRCases[LCOH_950],MATCH(D{lo_cap_row}&\"|\"&{FUEL},tblKBRCases[LookupKey],0))"
        f"-INDEX(tblKBRCases[LCOH_500],MATCH(D{lo_cap_row}&\"|\"&{FUEL},tblKBRCases[LookupKey],0)))/(950-500),\"N/A\"),"
        f"\"N/A\")",
        "Approximate cross-check at the lower bracket point only (illustrative, not a full re-derivation)",
    )
    put(
        "LCOH_CrossCheck_Delta_USD_per_kgH2",
        f'=IFERROR(D{lcoh_model_row}-D{lcoh_kbr_lookup_row},"N/A")',
        "Non-zero delta is expected (model uses continuous interpolation; KBR's own figure is a discrete table lookup)",
    )
    r += 1

    ws.cell(r, 2, "Duiker")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    duiker_opex_row = put(
        "Duiker_OPEX_MUSD_per_yr (base + license fees, converted to USD)",
        f"=IF({CAP}=12,"
        f"({nr.DUIKER_OPEX_EUR_M_PER_YR}+{nr.DUIKER_LICENSE_FEE_CONSTRUCTION_EUR}/1000000/25"
        f"+{nr.DUIKER_LICENSE_FEE_OPERATION_EUR_PER_T_H2}*{CAP}*1000/1000000)*{nr.EURUSD_FX_RATE},"
        f'"N/A -- Duiker package provides a single capacity point (12 ktpa)")',
        "Construction license fee (EUR2.7M) annualized straight-line over 25 yr; "
        "operation fee EUR8.50/t H2 x annual production, additive to base OPEX (EUR2.9M/yr)",
    )
    r += 1

    ws.cell(r, 2, "Casale / Topsoe / Technip")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    st.style_na_flag(ws.cell(r, 2, "Casale_OPEX_MUSD_per_yr"))
    st.style_na_flag(ws.cell(r, 4, "N/A -- technical proposal only, no OPEX data"))
    r += 1
    st.style_na_flag(ws.cell(r, 2, "Topsoe/Technip_OPEX_MUSD_per_yr"))
    st.style_na_flag(ws.cell(r, 4, "N/A -- no OPEX figure exists in any source in this repo"))
    r += 2

    ws.cell(r, 2, "Selected Output (Own & Operate mode)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "AnnualOPEX_Own (MUSD/yr)")
    st.style_label(ws.cell(r, 2))
    formula = (
        f'=IF({LIC}="KBR",D{kbr_opex_final_row},'
        f'IF({LIC}="Duiker",D{duiker_opex_row},'
        f'"N/A -- no OPEX data for this licensor"))'
    )
    ws.cell(r, 4, formula)
    st.style_calculated(ws.cell(r, 4))
    nr.register(wb, nr.ANNUAL_OPEX_OWN, f"'{nr.SHEET_CALC_OPEX}'!$D${r}")
    r += 1

    rows = {
        "kbr_opex_final_row": kbr_opex_final_row,
        "duiker_opex_row": duiker_opex_row,
        "in_range_row": in_range_row,
    }
    return ws, rows
