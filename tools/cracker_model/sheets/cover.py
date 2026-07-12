"""Cover sheet -- mdlguideline.md §2.1."""
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_COVER)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 90

    r = 2
    ws.cell(r, 2, "Ammonia Cracker Capacity Sizing & Project Economics Tool")
    st.style_title(ws.cell(r, 2))
    r += 1
    ws.cell(r, 2, "Version 1.0  |  Gentari Hydrogen Technical Center of Excellence")
    ws.cell(r, 2).font = st.body_font(italic=True)
    r += 2

    ws.cell(r, 2, "Business purpose")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    desc = (
        "Given a required hydrogen supply capacity, this tool estimates ammonia "
        "cracker sizing, CAPEX, OPEX, ISBL/OSBL cost, project unlevered IRR, and "
        "an indicative equipment list -- across five licensors (Topsoe, Technip, "
        "KBR, Casale, Duiker) and two commercial modes (Own & Operate a licensed "
        "unit, or pay a Tolling fee) -- to support a build-vs-tolling business "
        "decision. Every figure is either cited to a source document or "
        "explicitly labeled an ASSUMPTION; unavailable data is flagged N/A, "
        "never fabricated. See the Guide sheet for full scope and limitations."
    )
    ws.cell(r, 2, desc)
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 3, end_column=6)
    r += 5

    ws.cell(r, 2, "Engineering discipline")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, "Process / Chemical Engineering -- Ammonia Cracking & Hydrogen Production")
    r += 1
    ws.cell(r, 2, "Applicable standards")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, "ASME, API, ISO (equipment-level, where cited by source packages); "
                   "EU RFNBO (RED II/III); Korea KEEI Clean Hydrogen Certification")
    r += 2

    ws.cell(r, 2, "Version")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, "1.0")
    r += 1
    ws.cell(r, 2, "Release date")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, "=TODAY()")
    r += 1
    ws.cell(r, 2, "Author")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, "Gentari Hydrogen TCOE (AI-assisted build per mdlguideline.md)")
    r += 1
    ws.cell(r, 2, "Reviewer")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, "[Pending]")
    r += 1
    ws.cell(r, 2, "Approval status")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, "Draft -- for internal review")
    ws.cell(r, 4).font = st.body_font(bold=True, color=st.COLOR_WARNING_AMBER)
    r += 2

    ws.cell(r, 2, "Disclaimer")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    disclaimer = (
        "Licensor and tolling packages referenced in this workbook are commercially "
        "confidential and indicative (Class III-V cost estimates, +/-40% to +/-50% "
        "depending on source -- see the Constants sheet for the accuracy class carried "
        "with each figure). No figure in this workbook should be treated as firm "
        "pricing. Where a licensor's package does not state a value, this workbook "
        "shows an explicit N/A flag rather than an estimated or interpolated number, "
        "except where a labeled ASSUMPTION or a clearly-flagged regression "
        "extrapolation is used (see Guide sheet). This workbook does not assert "
        "RFNBO or KEEI certification status -- only whether an indicative CI figure "
        "would or would not clear the relevant threshold, subject to full "
        "verification."
    )
    ws.cell(r, 2, disclaimer)
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 5, end_column=6)
    r += 7

    ws.cell(r, 2, "Navigate")
    st.style_section_header(ws.cell(r, 2))
    r += 1
    nav_targets = [
        ("Start (Inputs)", nr.SHEET_INPUTS),
        ("User Guide", nr.SHEET_GUIDE),
        ("Dashboard", nr.SHEET_DASHBOARD),
        ("Comparison", nr.SHEET_COMPARISON),
        ("Report", nr.SHEET_REPORT),
    ]
    for label, target in nav_targets:
        cell = ws.cell(r, 2, f"-> {label}")
        cell.hyperlink = f"#'{target}'!A1"
        cell.font = Font(name=st.FONT_FAMILY, size=11, color="0563C1", underline="single")
        r += 1

    return ws
