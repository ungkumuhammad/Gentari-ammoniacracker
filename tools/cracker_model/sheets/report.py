"""Report -- printable A4 summary, formula-driven from Inputs/Dashboard/
Calc_ sheets. mdlguideline.md §2.6."""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_REPORT)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["D"].width = 40

    r = 2
    ws.cell(r, 2, "Ammonia Cracker Capacity Sizing & Project Economics -- Report")
    st.style_title(ws.cell(r, 2))
    r += 2

    def put(label, formula):
        nonlocal r
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        ws.cell(r, 4, formula)
        st.style_calculated(ws.cell(r, 4))
        r += 1

    ws.cell(r, 2, "Project Information")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    put("Project Name", "='Inputs'!$D$6")
    put("Client", "='Inputs'!$D$7")
    put("Date", "='Inputs'!$D$9")
    r += 1

    ws.cell(r, 2, "Input Summary")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    put("Licensor", f"={nr.LICENSOR_SELECTED}")
    put("Required H2 Capacity (ktpa)", f"={nr.REQUIRED_H2_CAPACITY_KTPA}")
    put("Cracker Fuel Mode", f"={nr.CRACKER_FUEL_MODE}")
    put("Commercial Mode", f"={nr.COMMERCIAL_MODE}")
    put("Project Region", f"={nr.PROJECT_REGION}")
    r += 1

    ws.cell(r, 2, "Results")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    put("CAPEX_Total_Own (MUSD)", f"={nr.CAPEX_TOTAL_OWN}")
    put("AnnualOPEX_Own (MUSD/yr)", f"={nr.ANNUAL_OPEX_OWN}")
    put("AnnualTollingFee_Tolling (MUSD/yr)", f"={nr.ANNUAL_TOLLING_FEE}")
    put("Unlevered IRR", f"={nr.IRR_RESULT}")
    put("NPV (MUSD)", f"={nr.NPV_RESULT}")
    r += 1

    ws.cell(r, 2, "Revision")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    ws.cell(r, 2, "Rev 1.0 -- Draft for internal review. See Settings for full version history.")
    ws.cell(r, 2).font = st.body_font(italic=True)
    r += 2

    ws.cell(r, 2, "Sign-off")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    for role in ["Prepared by", "Reviewed by", "Approved by"]:
        ws.cell(r, 2, role)
        st.style_label(ws.cell(r, 2), bold=False)
        ws.cell(r, 4, "")
        st.style_input(ws.cell(r, 4))
        r += 1

    return ws
