"""Constants sheet -- admin-protected, mdlguideline.md §2.3 color coding.

Holds every lookup table Calc_ sheets reference: tblKBRCapexPoints,
tblKBRCases (full capacity x fuel-mode grid), tblCasaleSchemes, Duiker/
Topsoe/Technip single-point tables, tblLicensors (master index),
tblTollingParties, tblRegionalFactors -- plus named single-cell constants
(regression fit, OSBL%, LHV, RFNBO/KEEI thresholds).
"""
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import data
from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

_TABLE_STYLE = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)


def _fact_cell(ws, row, col, fact: "data.Fact"):
    cell = ws.cell(row, col)
    if fact.value is None:
        cell.value = f"N/A -- {fact.na_reason}"
        st.style_na_flag(cell)
    elif fact.assumption is not None:
        cell.value = fact.value
        st.style_assumption_flag(cell)
    else:
        cell.value = fact.value
        st.style_calculated(cell)
    return cell


def _citation_text(obj) -> str:
    """Accepts a Fact, a Citation, or an Assumption and renders a display string."""
    c = obj
    if isinstance(obj, data.Fact):
        c = obj.citation or obj.assumption
    if isinstance(c, data.Citation):
        note = f" -- {c.note}" if c.note else ""
        return f"{c.doc}, {c.section} ({c.revision_date}){note}"
    if isinstance(c, data.Assumption):
        return f"ASSUMPTION: {c.note} [{c.memory_md_ref}]"
    return ""


def _add_table(ws, name, start_row, start_col, headers, rows):
    """Write a header row + data rows starting at (start_row, start_col),
    wrap as a structured Table, return the row after the table."""
    for j, h in enumerate(headers):
        c = ws.cell(start_row, start_col + j, h)
        st.style_label(c)
    r = start_row
    for row_vals in rows:
        r += 1
        for j, v in enumerate(row_vals):
            ws.cell(r, start_col + j, v)
    end_col_letter = get_column_letter(start_col + len(headers) - 1)
    start_col_letter = get_column_letter(start_col)
    ref = f"{start_col_letter}{start_row}:{end_col_letter}{r}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = _TABLE_STYLE
    ws.add_table(tbl)
    return r + 2


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_CONSTANTS)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col in "BCDEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 15
    ws.column_dimensions["Q"].width = 60

    r = 2
    ws.cell(r, 2, "Constants (Admin Protected)")
    st.style_title(ws.cell(r, 2))
    r += 1
    ws.cell(r, 2, "Password-protected. See Settings for the deterrent password. "
                   "Every figure below carries a citation or an ASSUMPTION label.")
    ws.cell(r, 2).font = st.body_font(italic=True)
    r += 2

    # --- tblKBRCapexPoints + interpolation/regression constants ---
    ws.cell(r, 2, "KBR ISBL CAPEX -- 4 Known Capacity Points (for interpolation & regression)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    kbr_rows = []
    for cap in data.KBR_CAPACITIES_KTPA:
        fact = data.KBR_ISBL_CAPEX_MUSD[cap]
        kbr_rows.append([cap, fact.value, _citation_text(fact)])
    kbr_pts_start = r  # header row of the table about to be written
    r = _add_table(ws, nr.TBL_KBR_CAPEX_POINTS, r, 2,
                    ["Capacity_ktpa", "ISBL_CAPEX_MUSD", "Citation"], kbr_rows)
    for i, cap in enumerate(data.KBR_CAPACITIES_KTPA):
        row_ref = kbr_pts_start + 1 + i
        nr.register(wb, [nr.KBR_CAP1, nr.KBR_CAP2, nr.KBR_CAP3, nr.KBR_CAP4][i],
                    f"'{nr.SHEET_CONSTANTS}'!$B${row_ref}")
        nr.register(wb, [nr.KBR_CAPEX1, nr.KBR_CAPEX2, nr.KBR_CAPEX3, nr.KBR_CAPEX4][i],
                    f"'{nr.SHEET_CONSTANTS}'!$C${row_ref}")

    ws.cell(r, 2, "KBR_OSBLPct_ASSUMPTION")
    st.style_label(ws.cell(r, 2))
    _fact_cell(ws, r, 4, data.KBR_OSBL_PCT_OF_ISBL)
    ws.cell(r, 6, _citation_text(data.KBR_OSBL_PCT_OF_ISBL))
    ws.cell(r, 6).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
    nr.register(wb, nr.KBR_OSBL_PCT_ASSUMPTION, f"'{nr.SHEET_CONSTANTS}'!$D${r}")
    r += 1

    ws.cell(r, 2, "KBR_LCOH_DiscountRate (%)")
    st.style_label(ws.cell(r, 2))
    _fact_cell(ws, r, 4, data.KBR_LCOH_DISCOUNT_RATE_PCT)
    nr.register(wb, nr.KBR_LCOH_DISCOUNT_RATE, f"'{nr.SHEET_CONSTANTS}'!$D${r}")
    r += 1

    ws.cell(r, 2, "KBR_ProjectLife_yr")
    st.style_label(ws.cell(r, 2))
    _fact_cell(ws, r, 4, data.KBR_PROJECT_LIFE_YR)
    nr.register(wb, nr.KBR_PROJECT_LIFE_YR, f"'{nr.SHEET_CONSTANTS}'!$D${r}")
    r += 1

    ws.cell(r, 2, "Casale_ProjectLife_yr")
    st.style_label(ws.cell(r, 2))
    _fact_cell(ws, r, 4, data.CASALE_PROJECT_LIFE_YR)
    nr.register(wb, nr.CASALE_PROJECT_LIFE_YR, f"'{nr.SHEET_CONSTANTS}'!$D${r}")
    r += 2

    ws.cell(r, 2, "Global log-log regression (extrapolation beyond 12-80 ktpa)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    cap_rng = f"'{nr.SHEET_CONSTANTS}'!$B${kbr_pts_start + 1}:$B${kbr_pts_start + 4}"
    capex_rng = f"'{nr.SHEET_CONSTANTS}'!$C${kbr_pts_start + 1}:$C${kbr_pts_start + 4}"
    ws.cell(r, 2, "CapacityRegressionSlope_b")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"=SLOPE(LN({capex_rng}),LN({cap_rng}))")
    st.style_calculated(ws.cell(r, 4))
    nr.register(wb, nr.CAPACITY_REGRESSION_SLOPE_B, f"'{nr.SHEET_CONSTANTS}'!$D${r}")
    r += 1
    ws.cell(r, 2, "CapacityRegressionIntercept_a")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"=EXP(INTERCEPT(LN({capex_rng}),LN({cap_rng})))")
    st.style_calculated(ws.cell(r, 4))
    nr.register(wb, nr.CAPACITY_REGRESSION_INTERCEPT_A, f"'{nr.SHEET_CONSTANTS}'!$D${r}")
    r += 1
    ws.cell(r, 2, "CapacityRegressionRSquared")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"=RSQ(LN({capex_rng}),LN({cap_rng}))")
    st.style_calculated(ws.cell(r, 4))
    nr.register(wb, nr.CAPACITY_REGRESSION_RSQUARED, f"'{nr.SHEET_CONSTANTS}'!$D${r}")
    r += 2

    # --- tblKBRCases: full capacity x fuel-mode grid ---
    ws.cell(r, 2, "KBR Full Case Grid (all capacity x fuel-mode combinations)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    headers = ["Capacity_ktpa", "FuelMode", "LookupKey", "H2_TPD", "NH3_Ratio_t_per_t",
               "EnergyEff_pct", "H2Yield_pct", "CI_kgCO2_kgH2", "Electricity_kWh_tH2",
               "OPEX_500", "OPEX_950", "OPEX_1100", "LCOH_500", "LCOH_950", "LCOH_1100"]
    case_start = r
    for j, h in enumerate(headers):
        st.style_label(ws.cell(r, 2 + j))
        ws.cell(r, 2 + j, h)
    for case in data.KBR_CASES:
        r += 1
        vals = [case.capacity_ktpa, data.FUEL_MODE_LABELS[case.fuel_mode],
                f"{case.capacity_ktpa}|{data.FUEL_MODE_LABELS[case.fuel_mode]}",
                case.h2_production_tpd, case.nh3_ratio_t_per_t, case.energy_efficiency_pct,
                case.h2_yield_pct, case.ci_kgco2_per_kgh2, case.electricity_kwh_per_tH2,
                case.opex_500, case.opex_950, case.opex_1100,
                case.lcoh_500, case.lcoh_950, case.lcoh_1100]
        for j, v in enumerate(vals):
            ws.cell(r, 2 + j, v)
    end_col = get_column_letter(2 + len(headers) - 1)
    tbl = Table(displayName="tblKBRCases", ref=f"B{case_start}:{end_col}{r}")
    tbl.tableStyleInfo = _TABLE_STYLE
    ws.add_table(tbl)
    ws.cell(r + 1, 2, _citation_text(data.KBR_CASES_CITATION))
    ws.cell(r + 1, 2).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
    r += 3

    # --- tblCasaleSchemes ---
    ws.cell(r, 2, "Casale Scheme Table (capacity-invariant -- see Guide)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    casale_rows = []
    for row in data.CASALE_ROWS:
        mapped = data.FUEL_MODE_LABELS[data.CASALE_SCHEME_TO_FUEL_MODE[row.scheme]]
        casale_rows.append([
            data.CASALE_SCHEME_LABELS[row.scheme], mapped, row.nh3_ratio_t_per_t_upto,
            row.h2_molar_efficiency_pct_upto, row.energy_efficiency_pct_upto, row.ci_kgco2_per_kgh2,
        ])
    r = _add_table(ws, "tblCasaleSchemes", r, 2,
                    ["Scheme", "MappedFuelMode_ASSUMPTION", "NH3_Ratio_upto_t_per_t",
                     "H2MolarEff_upto_pct", "EnergyEff_upto_pct", "CI_kgCO2_kgH2"], casale_rows)
    ws.cell(r - 1, 2, _citation_text(data.CASALE_ROWS_CITATION))
    ws.cell(r - 1, 2).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
    ws.cell(r, 2, data.CASALE_MAPPING_ASSUMPTION.note)
    st.style_assumption_flag(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 45
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True, vertical="top")
    r += 2

    # --- Duiker (single point) ---
    ws.cell(r, 2, "Duiker AHC (single capacity point: 12 ktpa)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    duiker_fields = [
        ("CAPEX_Total_EUR_M (no ISBL/OSBL split)", data.DUIKER_CAPEX_TOTAL_EUR_M),
        ("OPEX_EUR_M_per_yr", data.DUIKER_OPEX_EUR_M_PER_YR),
        ("LicenseFee_Construction_EUR", data.DUIKER_LICENSE_FEE_CONSTRUCTION_EUR),
        ("LicenseFee_Operation_EUR_per_tH2", data.DUIKER_LICENSE_FEE_OPERATION_EUR_PER_T_H2),
        ("NH3_Ratio_NH3fired_t_per_t", data.DUIKER_NH3_FIRED_RATIO),
        ("NH3_Ratio_NGfired_t_per_t", data.DUIKER_NG_FIRED_RATIO),
        ("CI_NH3fired_kgCO2_kgH2", data.DUIKER_CI_NH3_FIRED),
        ("CI_NGfired_kgCO2_kgH2", data.DUIKER_CI_NG_FIRED),
        ("H2_Purity_pct", data.DUIKER_H2_PURITY_PCT),
        ("EnergyEfficiency_pct", data.DUIKER_ENERGY_EFFICIENCY_PCT),
        ("Footprint_m2", data.DUIKER_FOOTPRINT_M2),
        ("OutputPressure_barg", data.DUIKER_OUTPUT_PRESSURE_BARG),
        ("ProjectLife_yr", data.DUIKER_PROJECT_LIFE_YR),
    ]
    duiker_named_ranges = {
        "CAPEX_Total_EUR_M (no ISBL/OSBL split)": nr.DUIKER_CAPEX_TOTAL_EUR_M,
        "OPEX_EUR_M_per_yr": nr.DUIKER_OPEX_EUR_M_PER_YR,
        "LicenseFee_Construction_EUR": nr.DUIKER_LICENSE_FEE_CONSTRUCTION_EUR,
        "LicenseFee_Operation_EUR_per_tH2": nr.DUIKER_LICENSE_FEE_OPERATION_EUR_PER_T_H2,
        "NH3_Ratio_NH3fired_t_per_t": nr.DUIKER_NH3_RATIO_NH3FIRED,
        "NH3_Ratio_NGfired_t_per_t": nr.DUIKER_NH3_RATIO_NGFIRED,
        "CI_NH3fired_kgCO2_kgH2": nr.DUIKER_CI_NH3FIRED,
        "CI_NGfired_kgCO2_kgH2": nr.DUIKER_CI_NGFIRED,
    }
    duiker_start = r
    for label, fact in duiker_fields:
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        _fact_cell(ws, r, 4, fact)
        ws.cell(r, 6, _citation_text(fact))
        ws.cell(r, 6).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
        if label in duiker_named_ranges:
            nr.register(wb, duiker_named_ranges[label], f"'{nr.SHEET_CONSTANTS}'!$D${r}")
        r += 1
    ws.cell(r, 2, data.DUIKER_MEMORY_MD_RATIO_DISCREPANCY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 45
    st.style_assumption_flag(ws.cell(r, 2))
    r += 2

    # --- Topsoe & Technip ---
    ws.cell(r, 2, "Topsoe & Technip (single TCOE-table data point each)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    tt_headers = ["Licensor", "Capacity_ktpa", "ISBL_CAPEX_MUSD", "NH3_Ratio", "H2Purity_pct",
                  "EnergyEff_pct", "CI_kgCO2e_kgH2", "Footprint"]
    tt_start = r
    for j, h in enumerate(tt_headers):
        st.style_label(ws.cell(r, 2 + j))
        ws.cell(r, 2 + j, h)
    for label, capacity, capex, ratio, purity, eff, ci, fp in [
        ("Topsoe", data.TOPSOE_CAPACITY_KTPA, data.TOPSOE_ISBL_CAPEX_MUSD, data.TOPSOE_NH3_RATIO,
         data.TOPSOE_H2_PURITY_PCT, data.TOPSOE_ENERGY_EFFICIENCY_PCT, data.TOPSOE_CI, data.TOPSOE_FOOTPRINT),
        ("Technip", data.TECHNIP_CAPACITY_KTPA, data.TECHNIP_ISBL_CAPEX_MUSD, data.TECHNIP_NH3_RATIO,
         data.TECHNIP_H2_PURITY_PCT, data.TECHNIP_ENERGY_EFFICIENCY_PCT, data.TECHNIP_CI, data.TECHNIP_FOOTPRINT),
    ]:
        r += 1
        ws.cell(r, 2, label)
        ws.cell(r, 3, capacity)
        _fact_cell(ws, r, 4, capex)
        _fact_cell(ws, r, 5, ratio)
        _fact_cell(ws, r, 6, purity)
        _fact_cell(ws, r, 7, eff)
        _fact_cell(ws, r, 8, ci)
        if label == "Topsoe":
            nr.register(wb, nr.TOPSOE_ISBL_CAPEX_MUSD, f"'{nr.SHEET_CONSTANTS}'!$D${r}")
            nr.register(wb, nr.TOPSOE_NH3_RATIO, f"'{nr.SHEET_CONSTANTS}'!$E${r}")
            nr.register(wb, nr.TOPSOE_CI, f"'{nr.SHEET_CONSTANTS}'!$H${r}")
        elif label == "Technip":
            nr.register(wb, nr.TECHNIP_NH3_RATIO, f"'{nr.SHEET_CONSTANTS}'!$E${r}")
            nr.register(wb, nr.TECHNIP_CI, f"'{nr.SHEET_CONSTANTS}'!$H${r}")
        _fact_cell(ws, r, 9, fp)
    tbl = Table(displayName="tblTopsoeTechnip", ref=f"B{tt_start}:I{r}")
    tbl.tableStyleInfo = _TABLE_STYLE
    ws.add_table(tbl)
    r += 1
    ws.cell(r, 2, data.TOPSOE_NO_PACKAGE_NOTE)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True, vertical="top")
    st.style_na_flag(ws.cell(r, 2))
    r += 2

    # --- tblLicensors master index ---
    ws.cell(r, 2, "Licensor Master Index")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    lic_headers = ["Licensor", "CapacityDataPoints", "CapacityInvariant", "AccuracyClass",
                   "ProjectLife_yr", "CostBasisRegion"]
    lic_rows = [
        ["Topsoe", 1, "No", "Class V", "N/A -- not stated", "N/A -- not stated"],
        ["Technip", 0, "No", "N/A -- CAPEX not available", "N/A -- not stated", "N/A -- not stated"],
        ["KBR", 4, "No", "Class V +/-50%", data.KBR_PROJECT_LIFE_YR.value, "East Asia (KBR's own words)"],
        ["Casale", 0, "Yes -- see Guide", "N/A -- technical proposal only", data.CASALE_PROJECT_LIFE_YR.value, "N/A -- no cost data"],
        ["Duiker", 1, "No", "+/-40%", "N/A -- not stated (defaults to 25 if unspecified)", "N/A -- not stated"],
    ]
    r = _add_table(ws, nr.TBL_LICENSORS, r, 2, lic_headers, lic_rows)

    # --- tblTollingParties ---
    ws.cell(r, 2, "Tolling Parties")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    toll_headers = ["Party", "TariffLow", "TariffHigh", "Currency", "CapacityMin_ktpa",
                    "CapacityMax_ktpa", "AccuracyClass"]
    toll_rows = []
    for party in [data.TOLLING_VOPAK_LINDE, data.TOLLING_VTTI, data.TOLLING_HOEGH_EVI]:
        toll_rows.append([party.name, party.tariff_low, party.tariff_high, party.currency,
                           party.capacity_min_ktpa or "N/A", party.capacity_max_ktpa or "N/A",
                           party.accuracy_class])
    toll_start = r
    r = _add_table(ws, nr.TBL_TOLLING_PARTIES, r, 2, toll_headers, toll_rows)
    for i, party in enumerate([data.TOLLING_VOPAK_LINDE, data.TOLLING_VTTI, data.TOLLING_HOEGH_EVI]):
        row_ref = toll_start + 1 + i
        key = party.name.split(" ")[0].replace("&", "").replace(",", "")
        nr.register(wb, nr.TOLLING_TARIFF_LOW.format(party=key), f"'{nr.SHEET_CONSTANTS}'!$C${row_ref}")
        nr.register(wb, nr.TOLLING_TARIFF_HIGH.format(party=key), f"'{nr.SHEET_CONSTANTS}'!$D${row_ref}")
        nr.register(wb, nr.TOLLING_CAPACITY_MIN.format(party=key), f"'{nr.SHEET_CONSTANTS}'!$F${row_ref}")
    ws.cell(r, 2, data.TOLLING_NIPPON_SANSO_EXCLUDED_NOTE)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 45
    st.style_na_flag(ws.cell(r, 2))
    r += 2

    # --- tblRegionalFactors ---
    ws.cell(r, 2, "Regional CAPEX Factors")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    reg_rows = [[region, fact.value, fact.assumption.note] for region, fact in data.REGIONAL_FACTORS.items()]
    r = _add_table(ws, nr.TBL_REGIONAL_FACTORS, r, 2, ["Region", "Factor", "Status"], reg_rows)
    ws.cell(r, 2, data.KBR_COST_BASIS_REGION_NOTE)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 45
    st.style_reference(ws.cell(r, 2))
    r += 2

    # --- shared engineering constants ---
    ws.cell(r, 2, "Shared Engineering Constants")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    for label, name, fact in [
        ("LHV_H2_MJperkg", nr.LHV_H2_MJ_PER_KG, data.LHV_H2_MJ_PER_KG),
        ("RFNBOCeiling_gCO2eperMJ", nr.RFNBO_CEILING_GCO2E_PER_MJ, data.RFNBO_CI_CEILING_GCO2E_PER_MJ),
        ("KEEIThreshold_kgCO2perkgH2", nr.KEEI_THRESHOLD_KGCO2_PER_KGH2, data.KEEI_CI_THRESHOLD_KGCO2_PER_KGH2),
        ("EURUSD_FXRate (Constants copy)", None, data.EURUSD_FX_RATE),
    ]:
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        _fact_cell(ws, r, 4, fact)
        ws.cell(r, 6, _citation_text(fact))
        ws.cell(r, 6).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
        if name:
            nr.register(wb, name, f"'{nr.SHEET_CONSTANTS}'!$D${r}")
        r += 1

    return ws
