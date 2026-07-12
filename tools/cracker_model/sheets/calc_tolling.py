"""Calc_Tolling -- tariff lookup, currency conversion, capacity-constraint
warnings, and AnnualTollingFee_Tolling for Calc_CashFlow_IRR."""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import data
from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

PARTY = nr.TOLLING_PARTY
CAP = nr.REQUIRED_H2_CAPACITY_KTPA


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_CALC_TOLLING)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["F"].width = 55

    r = 2
    ws.cell(r, 2, "Calc_Tolling")
    st.style_title(ws.cell(r, 2))
    r += 1
    ws.cell(r, 2, "Tolling fee shown at the midpoint of each party's sourced tariff range. "
                   "Equation: AnnualTollingFee_MUSD = Tariff_USD_per_kg x Capacity_ktpa "
                   "(1 ktpa H2 x 1 USD/kg = 1,000,000 kg x 1 USD/kg = 1 MUSD)")
    ws.cell(r, 2).font = st.body_font(italic=True)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 2

    def put(label, formula, note=""):
        nonlocal r
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        c = ws.cell(r, 4, formula)
        st.style_calculated(c)
        if note:
            ws.cell(r, 6, note)
            ws.cell(r, 6).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
        this_row = r
        r += 1
        return this_row

    ws.cell(r, 2, "Selected Tolling Party -- Lookup")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1

    tariff_low_row = put("TariffLow", f"=INDEX(tblTollingParties[TariffLow],MATCH({PARTY},tblTollingParties[Party],0))")
    tariff_high_row = put("TariffHigh", f"=INDEX(tblTollingParties[TariffHigh],MATCH({PARTY},tblTollingParties[Party],0))")
    currency_row = put("Currency", f"=INDEX(tblTollingParties[Currency],MATCH({PARTY},tblTollingParties[Party],0))")
    cap_min_row = put("CapacityMin_ktpa", f"=INDEX(tblTollingParties[CapacityMin_ktpa],MATCH({PARTY},tblTollingParties[Party],0))")
    r += 1

    ws.cell(r, 2, "Currency Conversion & Midpoint")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    low_usd_row = put(
        "TariffLow_USD_per_kg",
        f'=IF(D{currency_row}="EUR",D{tariff_low_row}*{nr.EURUSD_FX_RATE},D{tariff_low_row})',
    )
    high_usd_row = put(
        "TariffHigh_USD_per_kg",
        f'=IF(D{currency_row}="EUR",D{tariff_high_row}*{nr.EURUSD_FX_RATE},D{tariff_high_row})',
    )
    mid_usd_row = put(
        "TariffMidpoint_USD_per_kg (selected)",
        f"=AVERAGE(D{low_usd_row},D{high_usd_row})",
    )
    r += 1

    ws.cell(r, 2, "Capacity Constraint Check")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "CapacityWarning")
    st.style_label(ws.cell(r, 2))
    warn_cell = ws.cell(
        r, 4,
        f'=IF(ISNUMBER(D{cap_min_row}),IF({CAP}<D{cap_min_row},'
        f'"WARNING: "&{CAP}&" ktpa is below "&{PARTY}&"\'s minimum booking capacity ("&D{cap_min_row}&" ktpa)",""),"")',
    )
    st.style_assumption_flag(warn_cell)
    warn_row = r
    r += 2

    ws.cell(r, 2, "Passthrough Consumption (reference only, not separately costed in v1)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "Vopak & Linde")
    st.style_label(ws.cell(r, 2), bold=False)
    vopak_pt = data.TOLLING_VOPAK_PASSTHROUGH
    ws.cell(r, 4, f"NH3 feed {vopak_pt['ammonia_feed_t_per_t_h2'][0]}-{vopak_pt['ammonia_feed_t_per_t_h2'][1]} t/t H2; "
                   f"NH3 fuel {vopak_pt['ammonia_fuel_t_per_t_h2'][0]}-{vopak_pt['ammonia_fuel_t_per_t_h2'][1]} t/t H2; "
                   f"NG fuel {vopak_pt['ng_fuel_mwh_per_t_h2'][0]}-{vopak_pt['ng_fuel_mwh_per_t_h2'][1]} MWh/t H2; "
                   f"Power {vopak_pt['power_mw'][0]}-{vopak_pt['power_mw'][1]} MW")
    st.style_reference(ws.cell(r, 4))
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    r += 1
    ws.cell(r, 2, "VTTI")
    st.style_label(ws.cell(r, 2), bold=False)
    ws.cell(r, 4, data.TOLLING_VTTI_PASSTHROUGH_NOTE)
    st.style_reference(ws.cell(r, 4))
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    r += 2

    ws.cell(r, 2, "Selected Output (Tolling mode)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "AnnualTollingFee_MUSD_per_yr")
    st.style_label(ws.cell(r, 2))
    fee_cell = ws.cell(r, 4, f"=D{mid_usd_row}*{CAP}")
    st.style_calculated(fee_cell)
    nr.register(wb, nr.ANNUAL_TOLLING_FEE, f"'{nr.SHEET_CALC_TOLLING}'!$D${r}")
    r += 1

    rows = {"warn_row": warn_row}
    return ws, rows
