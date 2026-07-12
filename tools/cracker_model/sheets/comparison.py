"""Comparison -- all 5 licensor (Own & Operate) + 3 tolling-party rows,
side by side at the user's chosen capacity/fuel mode, independent of
whatever is currently selected on Inputs. Reuses each Calc_ sheet's
already-computed per-licensor cells directly (rather than re-deriving
formulas) to avoid duplicating -- and risking divergence from -- the
bracket/interpolation logic defined once in Calc_CapacitySizing/
Calc_OPEX/Calc_CarbonIntensity.

IRR shown for Own & Operate rows only, via RATE() (closed-form for a
constant annuity -- equivalent to the year-by-year IRR() in
Calc_CashFlow_IRR under the no-escalation assumption used throughout this
model, just far more compact than repeating a 30-row cash-flow table 8
times). Tolling rows have no CAPEX outlay, so IRR is mathematically
undefined (no sign change in the cash-flow stream) -- Annual Margin is
shown instead for a fair cross-mode comparison.
"""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

CAP = nr.REQUIRED_H2_CAPACITY_KTPA
REVENUE = f"({nr.OFFTAKE_H2_PRICE_USD_PER_KG}*{CAP})"

CS = nr.SHEET_CALC_CAPACITY_SIZING
CX = nr.SHEET_CALC_CAPEX_ISBL_OSBL
OX = nr.SHEET_CALC_OPEX
CI = nr.SHEET_CALC_CARBON_INTENSITY
TL = nr.SHEET_CALC_TOLLING


def build(wb, cs_rows, capex_rows, opex_rows, ci_rows) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_COMPARISON)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["I"].width = 55

    r = 2
    ws.cell(r, 2, "Comparison")
    st.style_title(ws.cell(r, 2))
    r += 1
    ws.cell(r, 2, "All licensor x mode permutations at the current Inputs capacity/fuel mode. "
                   "IRR via RATE() annuity formula (Own & Operate only; no escalation, matches "
                   "Calc_CashFlow_IRR's no-escalation basis).")
    ws.cell(r, 2).font = st.body_font(italic=True)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 2

    headers = ["Party / Mode", "CAPEX or Total Installed (MUSD)", "Annual OPEX or Fee (MUSD/yr)",
               "Direct CI (kgCO2/kgH2)", "Annual Margin (MUSD/yr)", "Unlevered IRR", "Notes"]
    header_row = r
    for j, h in enumerate(headers):
        c = ws.cell(r, 2 + j, h)
        st.style_label(c)
    r += 1

    def irr_formula(capex_ref, opex_ref):
        return (f'=IFERROR(RATE({nr.PROJECT_LIFE_YR},{REVENUE}-{opex_ref},-{capex_ref},0,0),'
                f'"N/A -- no real solution (check CAPEX/OPEX)")')

    def margin_formula(opex_ref):
        return f"={REVENUE}-{opex_ref}"

    own_operate_rows = [
        ("KBR (Own & Operate)", f"'{CX}'!$D${capex_rows['kbr_total_row']}", f"'{OX}'!$D${opex_rows['kbr_opex_final_row']}",
         f"'{CI}'!$D${ci_rows['kbr_ci_row']}"),
        ("Duiker (Own & Operate)", f"'{CX}'!$D${capex_rows['duiker_total_row']}", f"'{OX}'!$D${opex_rows['duiker_opex_row']}",
         f"'{CI}'!$D${ci_rows['duiker_ci_row']}"),
        ("Topsoe (Own & Operate)", f"'{CX}'!$D${capex_rows['topsoe_total_row']}", '"N/A -- no OPEX figure exists in any source"',
         f"'{CI}'!$D${ci_rows['topsoe_ci_row']}"),
        ("Casale (Own & Operate)", '"N/A -- technical proposal only"', '"N/A -- technical proposal only"',
         f"'{CI}'!$D${ci_rows['casale_ci_row']}"),
        ("Technip (Own & Operate)", '"N/A -- not available at current maturity"', '"N/A -- not available at current maturity"',
         f"'{CI}'!$D${ci_rows['technip_ci_row']}"),
    ]
    for label, capex_ref, opex_ref, ci_ref in own_operate_rows:
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        ws.cell(r, 3, f"={capex_ref}")
        st.style_calculated(ws.cell(r, 3))
        ws.cell(r, 4, f"={opex_ref}")
        st.style_calculated(ws.cell(r, 4))
        ws.cell(r, 5, f"={ci_ref}")
        st.style_calculated(ws.cell(r, 5))
        ws.cell(r, 6, margin_formula(f"D{r}"))
        st.style_calculated(ws.cell(r, 6))
        ws.cell(r, 7, irr_formula(f"C{r}", f"D{r}"))
        st.style_calculated(ws.cell(r, 7))
        r += 1

    tolling_parties = [
        ("Vopak & Linde (Tolling)", "Vopak & Linde (Antwerp)"),
        ("VTTI (Tolling)", "VTTI (Rotterdam/Antwerp, Amplifhy)"),
        ("Hoegh EVI (Tolling)", "Hoegh EVI (Floating Solution)"),
    ]
    for label, party_key in tolling_parties:
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        st.style_na_flag(ws.cell(r, 3, "N/A -- no CAPEX in Tolling mode"))
        low = f'INDEX(tblTollingParties[TariffLow],MATCH("{party_key}",tblTollingParties[Party],0))'
        high = f'INDEX(tblTollingParties[TariffHigh],MATCH("{party_key}",tblTollingParties[Party],0))'
        curr = f'INDEX(tblTollingParties[Currency],MATCH("{party_key}",tblTollingParties[Party],0))'
        mid_usd = (f'AVERAGE(IF({curr}="EUR",{low}*{nr.EURUSD_FX_RATE},{low}),'
                   f'IF({curr}="EUR",{high}*{nr.EURUSD_FX_RATE},{high}))')
        fee_formula = f"={mid_usd}*{CAP}"
        ws.cell(r, 4, fee_formula)
        st.style_calculated(ws.cell(r, 4))
        st.style_na_flag(ws.cell(r, 5, "See Calc_Tolling passthrough notes -- not separately costed"))
        ws.cell(r, 6, margin_formula(f"D{r}"))
        st.style_calculated(ws.cell(r, 6))
        st.style_na_flag(ws.cell(r, 7, "N/A -- IRR undefined (no CAPEX outlay); compare via Annual Margin"))
        r += 1

    tbl_end = r - 1
    from openpyxl.worksheet.table import Table, TableStyleInfo
    tbl = Table(displayName=nr.TBL_COMPARISON, ref=f"B{header_row}:H{tbl_end}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    return ws
