"""Calc_CapacitySizing -- ISBL CAPEX per licensor at the selected capacity.

KBR: piecewise log-log interpolation between the two bracketing known
points inside [12,80] ktpa (reproduces the 4 sourced values exactly);
global log-log regression outside that range, flagged as extrapolation.
Duiker/Topsoe: single data point, N/A elsewhere. Casale/Technip: always
N/A (no cost data in source package at any maturity).
"""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

CAP = f"{nr.REQUIRED_H2_CAPACITY_KTPA}"
LIC = f"{nr.LICENSOR_SELECTED}"


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_CALC_CAPACITY_SIZING)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["F"].width = 55

    r = 2
    ws.cell(r, 2, "Calc_CapacitySizing")
    st.style_title(ws.cell(r, 2))
    r += 1
    ws.cell(r, 2, "Equation: CAPEX(C) = Capex_lo * (C / Cap_lo) ^ b, "
                   "b = LN(Capex_hi/Capex_lo) / LN(Cap_hi/Cap_lo)")
    ws.cell(r, 2).font = st.body_font(italic=True)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 2

    # --- KBR piecewise interpolation ---
    ws.cell(r, 2, "KBR -- Bracket Selection & Piecewise Interpolation")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1

    def put(label, formula, note=""):
        nonlocal r
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        c = ws.cell(r, 4, formula)
        st.style_calculated(c)
        if note:
            ws.cell(r, 6, note)
            ws.cell(r, 6).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
        r += 1
        return c

    put("BracketLo_Cap",
        f"=IF({CAP}<={nr.KBR_CAP2},{nr.KBR_CAP1},IF({CAP}<={nr.KBR_CAP3},{nr.KBR_CAP2},{nr.KBR_CAP3}))")
    put("BracketLo_Capex",
        f"=IF({CAP}<={nr.KBR_CAP2},{nr.KBR_CAPEX1},IF({CAP}<={nr.KBR_CAP3},{nr.KBR_CAPEX2},{nr.KBR_CAPEX3}))")
    put("BracketHi_Cap",
        f"=IF({CAP}<={nr.KBR_CAP2},{nr.KBR_CAP2},IF({CAP}<={nr.KBR_CAP3},{nr.KBR_CAP3},{nr.KBR_CAP4}))")
    put("BracketHi_Capex",
        f"=IF({CAP}<={nr.KBR_CAP2},{nr.KBR_CAPEX2},IF({CAP}<={nr.KBR_CAP3},{nr.KBR_CAPEX3},{nr.KBR_CAPEX4}))")
    bracket_lo_cap_row, bracket_lo_capex_row = r - 4, r - 3
    bracket_hi_cap_row, bracket_hi_capex_row = r - 2, r - 1
    put("SegmentExponent_b",
        f"=LN(D{bracket_hi_capex_row}/D{bracket_lo_capex_row})/LN(D{bracket_hi_cap_row}/D{bracket_lo_cap_row})")
    segment_b_row = r - 1
    put("CAPEX_Piecewise_MUSD",
        f"=D{bracket_lo_capex_row}*({CAP}/D{bracket_lo_cap_row})^D{segment_b_row}",
        "Reproduces KBR's own 4 quoted values exactly at 12/24/68/80 ktpa")
    piecewise_row = r - 1
    put("CAPEX_Regression_MUSD",
        f"={nr.CAPACITY_REGRESSION_INTERCEPT_A}*{CAP}^{nr.CAPACITY_REGRESSION_SLOPE_B}",
        f"Global log-log fit, R2 = {nr.CAPACITY_REGRESSION_RSQUARED} (Constants sheet)")
    regression_row = r - 1
    put("InRange_12to80",
        f"=AND({CAP}>={nr.KBR_CAP1},{CAP}<={nr.KBR_CAP4})")
    in_range_row = r - 1
    ws.cell(r, 2, "KBR_ISBL_CAPEX_MUSD (final)")
    st.style_label(ws.cell(r, 2))
    kbr_capex_cell = ws.cell(r, 4, f"=IF(D{in_range_row},D{piecewise_row},D{regression_row})")
    st.style_calculated(kbr_capex_cell)
    kbr_capex_row = r
    r += 1
    ws.cell(r, 2, "KBR_Extrapolation_Flag")
    st.style_label(ws.cell(r, 2))
    extrap_cell = ws.cell(r, 4, f'=IF(D{in_range_row},"","REGRESSION EXTRAPOLATION -- beyond KBR\'s demonstrated 12-80 ktpa range, not vendor-confirmed, R2="&TEXT({nr.CAPACITY_REGRESSION_RSQUARED},"0.000"))')
    st.style_extrapolation_flag(extrap_cell)
    extrap_flag_row = r
    r += 2

    # --- Duiker / Topsoe: single point ---
    ws.cell(r, 2, "Duiker -- Single Capacity Point (12 ktpa)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "Duiker_CAPEX_Total_MUSD (converted from EUR, no ISBL/OSBL split)")
    st.style_label(ws.cell(r, 2))
    duiker_cell = ws.cell(r, 4,
        f'=IF({CAP}=12,{nr.DUIKER_CAPEX_TOTAL_EUR_M}*{nr.EURUSD_FX_RATE},'
        f'"N/A -- Duiker package provides a single capacity point (12 ktpa); no basis to scale or extrapolate")')
    st.style_calculated(duiker_cell)
    r += 1
    duiker_row = r - 1

    ws.cell(r, 2, "Topsoe -- Single Capacity Point (50 ktpa)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "Topsoe_ISBL_CAPEX_MUSD")
    st.style_label(ws.cell(r, 2))
    topsoe_cell = ws.cell(r, 4,
        f'=IF({CAP}=50,{nr.TOPSOE_ISBL_CAPEX_MUSD},'
        f'"N/A -- Topsoe package (TCOE synthesis table only) provides a single capacity point (50 ktpa); no basis to scale or extrapolate")')
    st.style_calculated(topsoe_cell)
    topsoe_row = r
    r += 1

    ws.cell(r, 2, "Casale / Technip -- No Cost Data at Any Capacity")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    casale_row = r
    ws.cell(r, 2, "Casale_CAPEX_MUSD")
    st.style_label(ws.cell(r, 2))
    na = ws.cell(r, 4, "N/A -- technical proposal only, no cost data (Rev.00, 05 Dec 2025)")
    st.style_na_flag(na)
    r += 1
    technip_row = r
    ws.cell(r, 2, "Technip_ISBL_CAPEX_MUSD")
    st.style_label(ws.cell(r, 2))
    na2 = ws.cell(r, 4, "N/A -- not available at current maturity (tcoedatabase §6.1)")
    st.style_na_flag(na2)
    r += 2

    # --- Selected outputs ---
    ws.cell(r, 2, "Selected Licensor Output (drives Calc_CAPEX_ISBL_OSBL)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "Selected_ISBL_CAPEX_Display")
    st.style_label(ws.cell(r, 2))
    selected_formula = (
        f'=IF({LIC}="KBR",D{kbr_capex_row},'
        f'IF({LIC}="Duiker",D{duiker_row},'
        f'IF({LIC}="Topsoe",D{topsoe_row},'
        f'IF({LIC}="Casale","N/A -- technical proposal only, no cost data",'
        f'"N/A -- not available at current maturity"))))'
    )
    sel_cell = ws.cell(r, 4, selected_formula)
    st.style_calculated(sel_cell)
    selected_capex_row = r
    r += 1
    ws.cell(r, 2, "Selected_Extrapolation_Flag")
    st.style_label(ws.cell(r, 2))
    sel_flag = ws.cell(r, 4, f'=IF({LIC}="KBR",D{extrap_flag_row},"")')
    st.style_extrapolation_flag(sel_flag)
    r += 1

    rows = {
        "kbr_capex_row": kbr_capex_row,
        "extrap_flag_row": extrap_flag_row,
        "duiker_row": duiker_row,
        "topsoe_row": topsoe_row,
        "selected_capex_row": selected_capex_row,
        "in_range_row": in_range_row,
    }
    return ws, rows
