"""Calc_CashFlow_IRR -- unlevered project cash flow, IRR, NPV.

Own & Operate: Year 0 = -CAPEX_Total_Own; Years 1..N = Capacity x
OfftakeH2Price - AnnualOPEX_Own. Tolling: Year 0 = 0; Years 1..N =
Capacity x OfftakeH2Price - AnnualTollingFee_Tolling. No debt/tax/
depreciation (explicitly out of scope, per the approved plan). A fixed
30-year cash-flow row range is used regardless of ProjectLife_yr --
years beyond ProjectLife_yr are zero-filled, which does not change the
IRR/NPV root-finding result (adding zero terms to a sum leaves it
unchanged), so this is a formula-generation convenience, not a modeling
simplification.
"""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

MODE = nr.COMMERCIAL_MODE
CAP = nr.REQUIRED_H2_CAPACITY_KTPA
LIC = nr.LICENSOR_SELECTED
MAX_YEARS = 30


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_CALC_CASHFLOW_IRR)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["F"].width = 55

    r = 2
    ws.cell(r, 2, "Calc_CashFlow_IRR")
    st.style_title(ws.cell(r, 2))
    r += 1
    ws.cell(r, 2, "Unlevered project cash flow -- no debt, tax, or depreciation modeled.")
    ws.cell(r, 2).font = st.body_font(italic=True)
    r += 2

    def put(label, formula, note="", style=st.style_calculated):
        nonlocal r
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        c = ws.cell(r, 4, formula)
        style(c)
        if note:
            ws.cell(r, 6, note)
            ws.cell(r, 6).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
        this_row = r
        r += 1
        return this_row

    ws.cell(r, 2, "Revenue & Annual Net Cash Flow")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    revenue_row = put(
        "Revenue_MUSD_per_yr (merchant sale)",
        f"={nr.OFFTAKE_H2_PRICE_USD_PER_KG}*{CAP}",
        "= USD/kg x ktpa (1 ktpa H2 = 1,000,000 kg -> result in MM USD)",
    )
    annual_ncf_row = put(
        "AnnualNetCashFlow_MUSD_per_yr",
        f'=IFERROR(IF({MODE}="Own & Operate",D{revenue_row}-{nr.ANNUAL_OPEX_OWN},D{revenue_row}-{nr.ANNUAL_TOLLING_FEE}),'
        f'"N/A -- cannot compute (see Calc_OPEX / Calc_Tolling for reason)")',
    )
    year0_row = put(
        "Year0_CashFlow_MUSD",
        f'=IFERROR(IF({MODE}="Own & Operate",-{nr.CAPEX_TOTAL_OWN},0),'
        f'"N/A -- cannot compute (see Calc_CAPEX_ISBL_OSBL for reason)")',
    )
    r += 1

    ws.cell(r, 2, "Project Life Basis Check")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    life_flag_row = put(
        "ProjectLifeMismatch_Flag",
        f'=IF(AND({LIC}="KBR",{nr.PROJECT_LIFE_YR}<>{nr.KBR_PROJECT_LIFE_YR}),'
        f'"NOTE: ProjectLife_yr ("&{nr.PROJECT_LIFE_YR}&") differs from KBR\'s own stated basis ("&{nr.KBR_PROJECT_LIFE_YR}&" yr)",'
        f'IF(AND({LIC}="Casale",{nr.PROJECT_LIFE_YR}<>{nr.CASALE_PROJECT_LIFE_YR}),'
        f'"NOTE: ProjectLife_yr ("&{nr.PROJECT_LIFE_YR}&") differs from Casale\'s own stated basis ("&{nr.CASALE_PROJECT_LIFE_YR}&" yr)",""))',
        style=st.style_assumption_flag,
    )
    r += 1

    ws.cell(r, 2, f"Cash Flow Table (Year 0-{MAX_YEARS}; years beyond ProjectLife_yr are zero-filled)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, "Year")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, "CashFlow_MUSD")
    st.style_label(ws.cell(r, 4))
    r += 1
    cf_start_row = r
    for year in range(0, MAX_YEARS + 1):
        ws.cell(r, 2, year)
        if year == 0:
            ws.cell(r, 4, f"=D{year0_row}")
        else:
            ws.cell(r, 4, f"=IF({year}<={nr.PROJECT_LIFE_YR},D{annual_ncf_row},0)")
        st.style_calculated(ws.cell(r, 4))
        r += 1
    cf_end_row = r - 1
    r += 1

    ws.cell(r, 2, "Results")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    cf_range = f"$D${cf_start_row}:$D${cf_end_row}"
    put(
        "IRR_Result (unlevered)",
        f'=IFERROR(IRR({cf_range}),"N/A -- cannot compute (check for N/A cash flow rows)")',
        f"IRR() over the full Year 0-{MAX_YEARS} range; trailing zero-filled years beyond "
        "ProjectLife_yr do not change the root-finding result",
    )
    irr_row = r - 1
    nr.register(wb, nr.IRR_RESULT, f"'{nr.SHEET_CALC_CASHFLOW_IRR}'!$D${irr_row}")
    put(
        "NPV_Result_MUSD (at DiscountRate_pct)",
        f'=IFERROR(D{year0_row}+NPV({nr.DISCOUNT_RATE_PCT}/100,$D${cf_start_row + 1}:$D${cf_end_row}),"N/A")',
        "Year 0 added un-discounted; NPV() discounts Year1..N starting at period 1",
    )
    npv_row = r - 1
    nr.register(wb, nr.NPV_RESULT, f"'{nr.SHEET_CALC_CASHFLOW_IRR}'!$D${npv_row}")

    rows = {"life_flag_row": life_flag_row, "irr_row": irr_row, "npv_row": npv_row}
    return ws, rows
