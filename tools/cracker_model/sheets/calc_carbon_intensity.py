"""Calc_CarbonIntensity -- direct CI per licensor/fuel-mode, unit
conversion, and RFNBO/KEEI screening (mdlguideline.md §9 wording rule:
never assert certification, only whether an indicative figure would/would
not clear the threshold)."""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import data
from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

LIC = nr.LICENSOR_SELECTED
CAP = nr.REQUIRED_H2_CAPACITY_KTPA
FUEL = nr.CRACKER_FUEL_MODE
CERT = nr.CERTIFICATION_FRAMEWORK
NG100 = "100% Natural Gas"
NG50 = "50% Natural Gas / 50% Ammonia (cracked gas)"
CF100 = "100% Clean Fuel (Ammonia / cracked gas)"


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_CALC_CARBON_INTENSITY)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["F"].width = 60

    r = 2
    ws.cell(r, 2, "Calc_CarbonIntensity")
    st.style_title(ws.cell(r, 2))
    r += 2

    def put(label, formula, note="", style=st.style_calculated):
        nonlocal r
        ws.cell(r, 2, label)
        st.style_label(ws.cell(r, 2), bold=False)
        c = ws.cell(r, 4, formula)
        style(c)
        if note:
            ws.cell(r, 6, note)
            ws.cell(r, 6).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
        this_row = r
        r += 1
        return this_row

    ws.cell(r, 2, "KBR -- CI Bracket Interpolation (linear, fuel-mode aware)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    is_ng100_row = put("FuelMode_Is_NG100", f'=({FUEL}="{NG100}")')
    lo_cap_row = put(
        "CI_BracketLo_Cap",
        f"=IF(D{is_ng100_row},IF({CAP}<={nr.KBR_CAP2},{nr.KBR_CAP1},IF({CAP}<={nr.KBR_CAP3},{nr.KBR_CAP2},{nr.KBR_CAP3})),{nr.KBR_CAP1})",
    )
    hi_cap_row = put(
        "CI_BracketHi_Cap",
        f"=IF(D{is_ng100_row},IF({CAP}<={nr.KBR_CAP2},{nr.KBR_CAP2},IF({CAP}<={nr.KBR_CAP3},{nr.KBR_CAP3},{nr.KBR_CAP4})),{nr.KBR_CAP4})",
    )
    key_lo_row = put("CI_LookupKey_Lo", f"=D{lo_cap_row}&\"|\"&{FUEL}")
    key_hi_row = put("CI_LookupKey_Hi", f"=D{hi_cap_row}&\"|\"&{FUEL}")
    ci_lo_row = put("CI_Lo_kgCO2_kgH2", f"=IFERROR(INDEX(tblKBRCases[CI_kgCO2_kgH2],MATCH(D{key_lo_row},tblKBRCases[LookupKey],0)),\"N/A\")")
    ci_hi_row = put("CI_Hi_kgCO2_kgH2", f"=IFERROR(INDEX(tblKBRCases[CI_kgCO2_kgH2],MATCH(D{key_hi_row},tblKBRCases[LookupKey],0)),\"N/A\")")
    kbr_ci_row = put(
        "KBR_CI_kgCO2_kgH2 (interpolated, linear)",
        f"=IF(D{lo_cap_row}=D{hi_cap_row},D{ci_lo_row},"
        f"D{ci_lo_row}+({CAP}-D{lo_cap_row})*(D{ci_hi_row}-D{ci_lo_row})/(D{hi_cap_row}-D{lo_cap_row}))",
        "Linear (not log-log) -- CI can be exactly 0 at Clean Fuel mode, which log-log cannot handle",
    )
    r += 1

    ws.cell(r, 2, "Casale -- Scheme Mapping (ASSUMPTION)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 2, data.CASALE_MAPPING_ASSUMPTION.note)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 60
    st.style_assumption_flag(ws.cell(r, 2))
    r += 1
    casale_ci_row = put(
        "Casale_CI_kgCO2_kgH2 (via mapped scheme)",
        f'=IF({FUEL}="{CF100}",INDEX(tblCasaleSchemes[CI_kgCO2_kgH2],MATCH("Self-sustaining (CO2-free)",tblCasaleSchemes[Scheme],0)),'
        f'IF({FUEL}="{NG50}",INDEX(tblCasaleSchemes[CI_kgCO2_kgH2],MATCH("Low-carbon (limited external fuel)",tblCasaleSchemes[Scheme],0)),'
        f'INDEX(tblCasaleSchemes[CI_kgCO2_kgH2],MATCH("Maximum fuel source",tblCasaleSchemes[Scheme],0))))',
        "Capacity-invariant -- same regardless of RequiredH2Capacity_ktpa",
        style=st.style_assumption_flag,
    )
    r += 1

    ws.cell(r, 2, "Duiker -- Two Discrete Firing Cases (no 50/50 blend data)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    duiker_ci_row = put(
        "Duiker_CI_kgCO2_kgH2",
        f'=IF({FUEL}="{CF100}",{nr.DUIKER_CI_NH3FIRED},IF({FUEL}="{NG100}",{nr.DUIKER_CI_NGFIRED},'
        f'"N/A -- Duiker package has no 50/50 NH3-NG blend CI data"))',
    )
    r += 1

    ws.cell(r, 2, "Topsoe / Technip -- Flat Figure (fuel-mode invariant in source)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    topsoe_ci_row = put("Topsoe_CI_kgCO2e_kgH2", f"={nr.TOPSOE_CI}")
    technip_ci_row = put("Technip_CI_kgCO2e_kgH2", f"={nr.TECHNIP_CI}")
    r += 1

    ws.cell(r, 2, "Selected CI & Unit Conversion")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    selected_ci_row = put(
        "Selected_CI_kgCO2_kgH2",
        f'=IF({LIC}="KBR",D{kbr_ci_row},IF({LIC}="Casale",D{casale_ci_row},'
        f'IF({LIC}="Duiker",D{duiker_ci_row},IF({LIC}="Topsoe",D{topsoe_ci_row},D{technip_ci_row}))))',
    )
    ci_gco2_mj_row = put(
        "Selected_CI_gCO2e_per_MJ",
        f"=IFERROR(D{selected_ci_row}*1000/{nr.LHV_H2_MJ_PER_KG},\"N/A\")",
        f"= kgCO2/kgH2 * 1000 g/kg / {nr.LHV_H2_MJ_PER_KG} MJ/kg (LHV H2)",
    )
    r += 1

    ws.cell(r, 2, "RFNBO / KEEI Screening (mdlguideline.md §9 wording -- never asserts certification)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    rfnbo_row = put(
        "RFNBO_Screening",
        f'=IFERROR(IF(D{ci_gco2_mj_row}<={nr.RFNBO_CEILING_GCO2E_PER_MJ},'
        f'"Based on the selected source\'s indicative CI of "&TEXT(D{ci_gco2_mj_row},"0.0")&" gCO2e/MJ, this WOULD clear the EU RFNBO ceiling ('
        f'"&{nr.RFNBO_CEILING_GCO2E_PER_MJ}&" gCO2e/MJ), subject to full additionality/correlation/certification verification.",'
        f'"Based on the selected source\'s indicative CI of "&TEXT(D{ci_gco2_mj_row},"0.0")&" gCO2e/MJ, this would NOT clear the EU RFNBO ceiling ('
        f'"&{nr.RFNBO_CEILING_GCO2E_PER_MJ}&" gCO2e/MJ), subject to full additionality/correlation/certification verification."),"N/A")',
    )
    ws.row_dimensions[rfnbo_row].height = 45
    keei_row = put(
        "KEEI_Screening",
        f'=IFERROR(IF(D{selected_ci_row}<{nr.KEEI_THRESHOLD_KGCO2_PER_KGH2},'
        f'"Based on the selected source\'s indicative CI of "&TEXT(D{selected_ci_row},"0.00")&" kgCO2/kgH2, this WOULD clear the Korea KEEI threshold ('
        f'"&{nr.KEEI_THRESHOLD_KGCO2_PER_KGH2}&" kgCO2/kgH2), subject to full certification verification.",'
        f'"Based on the selected source\'s indicative CI of "&TEXT(D{selected_ci_row},"0.00")&" kgCO2/kgH2, this would NOT clear the Korea KEEI threshold ('
        f'"&{nr.KEEI_THRESHOLD_KGCO2_PER_KGH2}&" kgCO2/kgH2), subject to full certification verification."),"N/A")',
    )
    ws.row_dimensions[keei_row].height = 45
    r += 1

    rows = {
        "lo_cap_row": lo_cap_row, "hi_cap_row": hi_cap_row,
        "key_lo_row": key_lo_row, "key_hi_row": key_hi_row,
        "selected_ci_row": selected_ci_row, "ci_gco2_mj_row": ci_gco2_mj_row,
        "rfnbo_row": rfnbo_row, "keei_row": keei_row,
        "kbr_ci_row": kbr_ci_row, "casale_ci_row": casale_ci_row,
        "duiker_ci_row": duiker_ci_row, "topsoe_ci_row": topsoe_ci_row,
        "technip_ci_row": technip_ci_row,
    }
    return ws, rows
