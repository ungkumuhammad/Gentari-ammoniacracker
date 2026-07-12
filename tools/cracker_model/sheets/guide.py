"""Guide sheet -- mdlguideline.md §2.2."""
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import data
from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st


def _section(ws, r, title, lines):
    ws.cell(r, 2, title)
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    for line in lines:
        ws.cell(r, 2, line)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 28
        r += 1
    return r + 1


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_GUIDE)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16

    r = 2
    ws.cell(r, 2, "User Guide")
    st.style_title(ws.cell(r, 2))
    r += 2

    r = _section(ws, r, "Purpose & scope", [
        "Estimate ammonia cracker capacity sizing, CAPEX/OPEX/ISBL/OSBL, an "
        "indicative equipment list, and unlevered project IRR, across 5 "
        "licensors and 2 commercial modes (Own & Operate / Tolling), to "
        "support a Gentari business decision on how to meet a required "
        "hydrogen supply capacity.",
    ])
    r = _section(ws, r, "Assumptions & limitations (read before using results)", [
        "Only KBR has a multi-point (4 capacities) CAPEX/OPEX dataset. Duiker "
        "and Topsoe have one data point each. Casale is technical-only (zero "
        "CAPEX/OPEX at any capacity). Technip has no CAPEX at any maturity.",
        f"KBR's H2 delivery pressure is stated inconsistently in its own source "
        f"package: {data.KBR_H2_DELIVERY_PRESSURE_CONFLICT}",
        "KBR's ISBL CAPEX basis is stated as 'similar technology plant in East "
        "Asia' (its own words) -- not Europe. See Calc_RegionalFactor.",
        "OSBL cost for KBR is known only at 12 ktpa (~55% of ISBL). Applied "
        "flat to 24/68/80 ktpa as a labeled ASSUMPTION -- KBR's own package "
        "states this % declines at larger scale but gives no second point.",
        "Casale's fuel-mode figures are mapped onto the shared NG100/NG50/"
        "CF100 dropdown as a labeled ASSUMPTION -- Casale's own package never "
        "states this correspondence. See Calc_CarbonIntensity.",
        "Capacities outside KBR's demonstrated 12-80 ktpa range use a global "
        "log-log regression across KBR's 4 known points, visually flagged "
        "orange as REGRESSION EXTRAPOLATION -- not a vendor-confirmed figure.",
        "No IRR, discount rate, or payback figure exists in any source "
        "document in this repo. DiscountRate_pct and OfftakeH2Price_USDperkg "
        "are pure user ASSUMPTIONS (yellow-filled on Inputs).",
        "EUR-denominated figures (Duiker, Vopak & Linde, VTTI) are converted "
        "via EURUSD_FXRate, a snapshot ECB reference rate -- see Constants.",
        "No debt, tax, or depreciation schedule is modeled -- IRR is "
        "unlevered (CAPEX or tolling-fee outflow vs. net operating cash "
        "flow only).",
    ])
    r = _section(ws, r, "Required inputs & units", [
        "All inputs are entered on the Inputs sheet only (light yellow = "
        "assumption/free entry, light blue = dropdown). Capacity in ktpa H2 "
        "(100% H2 basis). Prices in USD unless noted. CAPEX/OPEX in MM USD. "
        "Carbon intensity shown in both kgCO2/kgH2 and gCO2e/MJ (LHV H2 = "
        "120 MJ/kg, per mdlguideline.md §9).",
    ])
    r = _section(ws, r, "Calculation workflow", [
        "Inputs -> Constants lookup -> Calc_CapacitySizing -> "
        "[Calc_CAPEX_ISBL_OSBL + Calc_OPEX] (Own & Operate) or Calc_Tolling "
        "(Tolling) -> Calc_RegionalFactor -> Calc_CarbonIntensity -> "
        "Calc_CashFlow_IRR -> Dashboard -> Comparison -> Report.",
    ])
    r = _section(ws, r, "Output interpretation", [
        "N/A (red) means the selected licensor's source package does not "
        "provide that figure -- never treat as zero. ASSUMPTION (amber) "
        "means a labeled, admin-editable, non-sourced modeling input. "
        "REGRESSION EXTRAPOLATION (orange) means the value is outside "
        "KBR's vendor-demonstrated capacity range.",
        "RFNBO/KEEI screening never asserts certification -- it states "
        "whether an indicative CI figure would or would not clear the "
        "threshold, subject to full additionality/correlation/certification "
        "verification, per CLAUDE.md §4.3 and mdlguideline.md §9.",
    ])
    r = _section(ws, r, "References", [
        "Licensor/kbr/kbr-johor-hub.md (KBR H2ACT, Rev 0, 23 Dec 2025)",
        "Licensor/Casale/ (MACH2 Technical Proposal, Design Basis, Process "
        "Description, Rev.00, 05 Dec 2025)",
        "Licensor/duiker/duiker-johor-hub.md (AHC Budgetary Proposal 122380, "
        "05 Dec 2025)",
        "tcoedatabase/WIP_Ammonia_Cracker_Database.md (Rev 30-Jun-2026)",
        "tolling/vopak/Vopak_Cracker_Tolling_Fee.md (LoI, 20 Aug 2025)",
        "tolling/vtti/VTTI_Cracker_Tolling_Fee.md (Process Letter 2, 25 Apr 2025)",
        "ECB Euro foreign exchange reference rates (EURUSD_FXRate default, "
        "3 Jul 2026)",
        "CLAUDE.md (No-Fabrication Rule, RFNBO/KEEI framework) and "
        "mdlguideline.md (workbook standard) -- both at repo root",
    ])
    r = _section(ws, r, "FAQ", [
        "Q: Why does changing Licensor to Casale zero out CAPEX/OPEX? "
        "A: Casale's Dec-2025 package is technical-only -- no cost data "
        "exists for Casale at any capacity in this repo.",
        "Q: Why is Duiker locked to ~12 ktpa? A: Duiker's package provides "
        "exactly one capacity point -- there is no basis to scale it.",
        "Q: Can I turn off the OSBL 55% assumption? A: Yes, edit "
        "KBR_OSBLPct_ASSUMPTION on the (password-protected) Constants sheet.",
    ])

    return ws
