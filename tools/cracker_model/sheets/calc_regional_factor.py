"""Calc_RegionalFactor -- CAPEX multiplier lookup, Own & Operate mode only."""
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import data
from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

REGION = nr.PROJECT_REGION


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_CALC_REGIONAL_FACTOR)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["F"].width = 60

    r = 2
    ws.cell(r, 2, "Calc_RegionalFactor")
    st.style_title(ws.cell(r, 2))
    r += 2

    ws.cell(r, 2, "RegionalFactorSelected")
    st.style_label(ws.cell(r, 2))
    cell = ws.cell(r, 4, f"=INDEX(tblRegionalFactors[Factor],MATCH({REGION},tblRegionalFactors[Region],0))")
    st.style_assumption_flag(cell)
    nr.register(wb, nr.REGIONAL_FACTOR_SELECTED, f"'{nr.SHEET_CALC_REGIONAL_FACTOR}'!$D${r}")
    r += 1
    ws.cell(r, 2, "RegionalFactorStatus")
    st.style_label(ws.cell(r, 2))
    ws.cell(r, 4, f"=INDEX(tblRegionalFactors[Status],MATCH({REGION},tblRegionalFactors[Region],0))")
    st.style_assumption_flag(ws.cell(r, 4))
    status_row = r
    r += 2

    ws.cell(r, 2, "KBR Cost Basis vs. Selected Region")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    ws.cell(r, 2, data.KBR_COST_BASIS_REGION_NOTE)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 60
    st.style_reference(ws.cell(r, 2))
    r += 1
    ws.cell(r, 2, data.REGIONAL_FACTOR_NO_SOURCE_NOTE)
    ws.cell(r, 2).alignment = st.Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 45
    st.style_assumption_flag(ws.cell(r, 2))
    r += 1

    rows = {"status_row": status_row}
    return ws, rows
