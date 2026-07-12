"""Inputs sheet -- the only user-editable sheet. mdlguideline.md §2.3."""
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from tools.cracker_model import data
from tools.cracker_model import named_ranges as nr
from tools.cracker_model import styles as st

LICENSORS = ["Topsoe", "Technip", "KBR", "Casale", "Duiker"]
COMMERCIAL_MODES = ["Own & Operate", "Tolling"]
TOLLING_PARTIES = ["Vopak & Linde", "VTTI", "Hoegh EVI"]
FUEL_MODE_CHOICES = [data.FUEL_MODE_LABELS[m] for m in data.FUEL_MODES]
CERTIFICATION_FRAMEWORKS = ["EU RFNBO", "Korea KEEI", "Both"]
REGIONS = list(data.REGIONAL_FACTORS.keys())


def _add_dv(ws, formula1, cell_ref):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=False, showDropDown=False)
    dv.error = "Please choose a value from the dropdown list."
    dv.errorTitle = "Invalid selection"
    ws.add_data_validation(dv)
    dv.add(cell_ref)
    return dv


def _row(ws, r, label, value, cell_col="D", note=None, dropdown=False, input_cell=True, name=None):
    ws.cell(r, 2, label)
    st.style_label(ws.cell(r, 2), bold=False)
    col_idx = {"D": 4}[cell_col]
    cell = ws.cell(r, col_idx, value)
    if dropdown:
        st.style_dropdown(cell, locked=False)
    elif input_cell:
        st.style_input(cell, locked=False)
    else:
        st.style_calculated(cell)
    if note:
        ws.cell(r, col_idx + 2, note)
        ws.cell(r, col_idx + 2).font = st.body_font(italic=True, color=st.COLOR_SECONDARY_STEEL_GRAY)
    if name:
        nr.register(ws.parent, name, f"'{nr.SHEET_INPUTS}'!${cell_col}${r}")
    return cell


def build(wb) -> Worksheet:
    ws = wb.create_sheet(nr.SHEET_INPUTS)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["F"].width = 46

    r = 2
    ws.cell(r, 2, "Inputs")
    st.style_title(ws.cell(r, 2))
    r += 1
    ws.cell(r, 2, "Light yellow = free-entry assumption. Light blue = dropdown. Edit only cells in this sheet.")
    ws.cell(r, 2).font = st.body_font(italic=True)
    r += 2

    # --- Project Information ---
    ws.cell(r, 2, "Project Information")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    _row(ws, r, "Project Name", "Ammonia Cracker Sizing Study"); r += 1
    _row(ws, r, "Client", "Gentari Hydrogen"); r += 1
    _row(ws, r, "Engineer", "[Name]"); r += 1
    _row(ws, r, "Date", "=TODAY()"); r += 2

    # --- Cracker / Licensor Parameters ---
    ws.cell(r, 2, "Cracker / Licensor Parameters")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    lic_cell = _row(ws, r, "Licensor", "KBR", dropdown=True, name=nr.LICENSOR_SELECTED,
                     note="Only KBR has a 4-point capacity/CAPEX dataset -- see Guide")
    _add_dv(ws, '"' + ",".join(LICENSORS) + '"', lic_cell.coordinate)
    r += 1
    cap_cell = _row(ws, r, "Required H2 Capacity (ktpa)", 12, name=nr.REQUIRED_H2_CAPACITY_KTPA,
                     note="KBR piecewise fit valid 12-80 ktpa; outside that, regression extrapolation applies")
    dv_cap = DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=False)
    dv_cap.error = "Please enter a capacity greater than zero."
    dv_cap.errorTitle = "Invalid Input"
    dv_cap.promptTitle = "Required H2 Capacity"
    dv_cap.prompt = "Enter required hydrogen supply capacity in ktpa (100% H2 basis)."
    dv_cap.showInputMessage = True
    ws.add_data_validation(dv_cap)
    dv_cap.add(cap_cell.coordinate)
    r += 1
    fuel_cell = _row(ws, r, "Cracker Fuel Mode", FUEL_MODE_CHOICES[0], dropdown=True,
                      name=nr.CRACKER_FUEL_MODE,
                      note="KBR: full data at 12 & 80 ktpa only; 24/68 ktpa are NG-only")
    _add_dv(ws, '"' + ",".join(FUEL_MODE_CHOICES) + '"', fuel_cell.coordinate)
    r += 2

    # --- Commercial Mode ---
    ws.cell(r, 2, "Commercial Mode")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    mode_cell = _row(ws, r, "Commercial Mode", COMMERCIAL_MODES[0], dropdown=True,
                      name=nr.COMMERCIAL_MODE)
    _add_dv(ws, '"' + ",".join(COMMERCIAL_MODES) + '"', mode_cell.coordinate)
    r += 1
    toll_cell = _row(ws, r, "Tolling Party", TOLLING_PARTIES[0], dropdown=True,
                      name=nr.TOLLING_PARTY,
                      note="(used only when Commercial Mode = Tolling)")
    _add_dv(ws, '"' + ",".join(TOLLING_PARTIES) + '"', toll_cell.coordinate)
    r += 2

    # --- Certification Screening ---
    ws.cell(r, 2, "Certification Screening")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    cert_cell = _row(ws, r, "Target Framework", CERTIFICATION_FRAMEWORKS[0], dropdown=True,
                      name=nr.CERTIFICATION_FRAMEWORK)
    _add_dv(ws, '"' + ",".join(CERTIFICATION_FRAMEWORKS) + '"', cert_cell.coordinate)
    r += 2

    # --- Regional Factor ---
    ws.cell(r, 2, "Regional Factor")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    region_cell = _row(ws, r, "Project Region", REGIONS[0], dropdown=True,
                        name=nr.PROJECT_REGION,
                        note="Applies a CAPEX multiplier -- see Calc_RegionalFactor; all factors default to 1.00, no citable free source")
    _add_dv(ws, '"' + ",".join(REGIONS) + '"', region_cell.coordinate)
    r += 2

    # --- Economics (ASSUMPTION) ---
    ws.cell(r, 2, "Economics (ASSUMPTION -- yellow, admin/user-editable)")
    st.style_section_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    _row(ws, r, "Offtake H2 Price (USD/kg)", 5.00, name=nr.OFFTAKE_H2_PRICE_USD_PER_KG,
         note="Merchant sale price to customer -- pure ASSUMPTION, no market series in repo"); r += 1
    _row(ws, r, "Project Life (yr)", 25, name=nr.PROJECT_LIFE_YR,
         note="KBR default 25 yr; Casale's own basis is 20 yr -- Dashboard flags a mismatch"); r += 1
    _row(ws, r, "Discount Rate (%)", 10.0, name=nr.DISCOUNT_RATE_PCT,
         note="Gentari hurdle rate ASSUMPTION -- distinct from KBR's own 8% LCOH rate"); r += 1
    _row(ws, r, "Ammonia Price (USD/t)", 500, name=nr.AMMONIA_PRICE_USD_PER_T,
         note="KBR OPEX interpolated between its $500/$950/$1100 per t data points"); r += 1
    _row(ws, r, "Electricity Price (USD/MWh)", 140, name=nr.ELECTRICITY_PRICE_USD_PER_MWH,
         note="KBR's own cost-input assumption, reused (no separate Gentari figure sourced)"); r += 1
    _row(ws, r, "Natural Gas Price (USD/MWh)", 30, name=nr.NG_PRICE_USD_PER_MWH,
         note="KBR's own cost-input assumption"); r += 1
    _row(ws, r, "EUR/USD FX Rate", data.EURUSD_FX_RATE.value, name=nr.EURUSD_FX_RATE,
         note="ECB reference rate, 3 Jul 2026 snapshot -- admin-editable"); r += 1

    ws.sheet_properties.tabColor = st.COLOR_PRIMARY_DEEP_BLUE
    return ws
