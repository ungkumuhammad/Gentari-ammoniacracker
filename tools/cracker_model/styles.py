"""mdlguideline.md §4-6 typography/color constants and cell-style helpers.

Font family: Calibri is used (not Segoe UI) for cross-platform reliability
-- this build's QA pipeline recalculates/renders via headless LibreOffice,
where Segoe UI is not guaranteed to be installed; Calibri is in
mdlguideline.md's own preferred-fonts list (§4) and renders identically
across Excel/LibreOffice.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.protection import Protection

# --- mdlguideline.md §6 cell-type fills ---
FILL_USER_INPUT = "FFF2CC"
FILL_DROPDOWN = "DCE6F1"
FILL_CALCULATED = "F2F2F2"
FILL_REFERENCE = "FFFFFF"
FILL_ERROR = "F8CBAD"
FILL_HEADER = "1F3864"
FILL_SUCCESS = "C6E0B4"
FILL_WARNING = "FFE699"
FILL_EXTRAPOLATION = "FCE4D6"   # orange-tinted, distinct from FILL_WARNING amber

# --- mdlguideline.md §5 theme colors ---
COLOR_PRIMARY_DEEP_BLUE = "1F3864"
COLOR_SECONDARY_STEEL_GRAY = "808080"
COLOR_ACCENT_ORANGE = "ED7D31"
COLOR_SUCCESS_GREEN = "375623"
COLOR_WARNING_AMBER = "BF8F00"
COLOR_ERROR_RED = "C00000"
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "000000"

FONT_FAMILY = "Calibri"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


FILL_INPUT_PATTERN = _fill(FILL_USER_INPUT)
FILL_DROPDOWN_PATTERN = _fill(FILL_DROPDOWN)
FILL_CALCULATED_PATTERN = _fill(FILL_CALCULATED)
FILL_REFERENCE_PATTERN = _fill(FILL_REFERENCE)
FILL_ERROR_PATTERN = _fill(FILL_ERROR)
FILL_HEADER_PATTERN = _fill(FILL_HEADER)
FILL_SUCCESS_PATTERN = _fill(FILL_SUCCESS)
FILL_WARNING_PATTERN = _fill(FILL_WARNING)
FILL_EXTRAPOLATION_PATTERN = _fill(FILL_EXTRAPOLATION)

THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF") for _ in range(4)))


def title_font() -> Font:
    return Font(name=FONT_FAMILY, size=18, bold=True, color=COLOR_PRIMARY_DEEP_BLUE)


def section_header_font() -> Font:
    return Font(name=FONT_FAMILY, size=13, bold=True, color=COLOR_WHITE)


def body_font(bold: bool = False, italic: bool = False, color: str = COLOR_BLACK) -> Font:
    return Font(name=FONT_FAMILY, size=11, bold=bold, italic=italic, color=color)


def style_title(cell) -> None:
    cell.font = title_font()
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_section_header(cell) -> None:
    cell.font = section_header_font()
    cell.fill = FILL_HEADER_PATTERN
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def style_input(cell, locked: bool = False) -> None:
    """User-editable ASSUMPTION/input cell -- light yellow, unlocked."""
    cell.fill = FILL_INPUT_PATTERN
    cell.font = body_font()
    cell.border = THIN_BORDER
    cell.protection = Protection(locked=locked)


def style_dropdown(cell, locked: bool = False) -> None:
    cell.fill = FILL_DROPDOWN_PATTERN
    cell.font = body_font()
    cell.border = THIN_BORDER
    cell.protection = Protection(locked=locked)


def style_calculated(cell) -> None:
    cell.fill = FILL_CALCULATED_PATTERN
    cell.font = body_font()
    cell.border = THIN_BORDER
    cell.protection = Protection(locked=True)


def style_reference(cell) -> None:
    cell.fill = FILL_REFERENCE_PATTERN
    cell.font = body_font(italic=True, color=COLOR_SECONDARY_STEEL_GRAY)
    cell.protection = Protection(locked=True)


def style_na_flag(cell) -> None:
    cell.fill = FILL_ERROR_PATTERN
    cell.font = body_font(italic=True, color=COLOR_ERROR_RED)
    cell.protection = Protection(locked=True)


def style_assumption_flag(cell) -> None:
    cell.fill = FILL_WARNING_PATTERN
    cell.font = body_font(italic=True, color=COLOR_WARNING_AMBER)
    cell.protection = Protection(locked=True)


def style_extrapolation_flag(cell) -> None:
    cell.fill = FILL_EXTRAPOLATION_PATTERN
    cell.font = body_font(bold=True, color=COLOR_ACCENT_ORANGE)
    cell.protection = Protection(locked=True)


def style_pass(cell) -> None:
    cell.fill = FILL_SUCCESS_PATTERN
    cell.font = body_font(bold=True, color=COLOR_SUCCESS_GREEN)


def style_label(cell, bold: bool = True) -> None:
    cell.font = body_font(bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="center")
